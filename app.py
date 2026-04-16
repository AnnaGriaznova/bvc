import os
import sys
import json
import logging
import asyncio
import random
import io
from typing import Any, Dict, Optional, List
from copy import deepcopy
from datetime import datetime

import aiohttp
from aiohttp import web
import asyncpg
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PIL import Image, ImageDraw, ImageFont

# VK Bot - Академия BVC v2 (multi-group, per-course tests)
sys.stdout.reconfigure(line_buffering=True)
sys.stderr.reconfigure(line_buffering=True)

# -----------------------------------------------------------------------------
# Environment Variables
# -----------------------------------------------------------------------------
USER_MEN_STR = os.getenv("USER_MEN", "")  # Manager IDs - can open access, add fortune wheel spins
USER_MAR_STR = os.getenv("USER_MAR", "")  # Marketing IDs - can view form answers
USER_ADMIN = os.getenv("USER_ADMIN", "")  # Super admin ID for admin panel and database export
PORT = int(os.getenv("PORT", "8080"))

# Database configuration (Amvera PostgreSQL)
DB_HOST = os.getenv("DB_HOST", "")
DB_NAME = os.getenv("DB_NAME", "")
DB_USER = os.getenv("DB_USER", "")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")

# Multi-group configuration
# Parse GROUP_X, TOKEN_X, CONFIRMATION_TOKEN_X from environment (X = 1, 2, 3, ...)
# Format: GROUP_CONFIGS = {group_id: {"token": "...", "confirmation_token": "..."}}
GROUP_CONFIGS: Dict[int, Dict[str, str]] = {}
GROUP_COURSE_MAP: Dict[int, int] = {}  # group_id -> course_index (1-based)
GROUP_INDEX_MAP: Dict[int, int] = {}  # group_id -> env index (1-based)
_group_index = 1
while True:
    _group_id_str = os.getenv(f"GROUP_{_group_index}", "")
    _token = os.getenv(f"TOKEN_{_group_index}", "")
    _confirm = os.getenv(f"CONFIRMATION_TOKEN_{_group_index}", "")
    if not _group_id_str or not _token or not _confirm:
        break
    try:
        _gid = int(_group_id_str)
        GROUP_CONFIGS[_gid] = {"token": _token, "confirmation_token": _confirm}
        GROUP_COURSE_MAP[_gid] = _group_index  # group_index IS the course number
        GROUP_INDEX_MAP[_gid] = _group_index
        print(f"Group {_gid} configured (TOKEN_{_group_index}, CONFIRMATION_TOKEN_{_group_index}) -> Course {_group_index}", flush=True)
    except ValueError:
        print(f"WARNING: Invalid GROUP_{_group_index} value '{_group_id_str}', skipping", flush=True)
    _group_index += 1

# For backward compatibility: if GROUP_1 not set, try old single-group env vars
if not GROUP_CONFIGS:
    _old_token = os.getenv("TOKEN", "")
    _old_confirm = os.getenv("CONFIRMATION_TOKEN", "")
    _old_group = os.getenv("GROUP_1", "")
    if _old_token and _old_confirm and _old_group:
        try:
            _gid = int(_old_group)
            GROUP_CONFIGS[_gid] = {"token": _old_token, "confirmation_token": _old_confirm}
            print(f"Group {_gid} configured (legacy TOKEN/CONFIRMATION_TOKEN)", flush=True)
        except ValueError:
            pass

# First group as default
DEFAULT_GROUP_ID = next(iter(GROUP_CONFIGS), 0)
DEFAULT_TOKEN = GROUP_CONFIGS[DEFAULT_GROUP_ID]["token"] if DEFAULT_GROUP_ID else ""
DEFAULT_CONFIRMATION_TOKEN = GROUP_CONFIGS[DEFAULT_GROUP_ID]["confirmation_token"] if DEFAULT_GROUP_ID else ""

# Parse manager IDs (USER_MEN) from comma-separated string
USER_MEN_IDS = []
for _id in USER_MEN_STR.split(","):
    _id = _id.strip()
    if _id:
        try:
            USER_MEN_IDS.append(int(_id))
        except ValueError:
            print(f"WARNING: Invalid manager ID '{_id}' in USER_MEN, skipping", flush=True)

# Parse marketing IDs (USER_MAR) from comma-separated string
USER_MAR_IDS = []
for _id in USER_MAR_STR.split(","):
    _id = _id.strip()
    if _id:
        try:
            USER_MAR_IDS.append(int(_id))
        except ValueError:
            print(f"WARNING: Invalid marketing ID '{_id}' in USER_MAR, skipping", flush=True)

# Parse super admin ID
try:
    USER_ADMIN_ID = int(USER_ADMIN) if USER_ADMIN else 0
except ValueError:
    print(f"WARNING: Invalid USER_ADMIN ID '{USER_ADMIN}', defaulting to 0", flush=True)
    USER_ADMIN_ID = 0

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger("vk_bot")

# Print startup info
print("=" * 50, flush=True)
print("VK BOT STARTING [v2.1-main]", flush=True)
print(f"GROUPS CONFIGURED: {len(GROUP_CONFIGS)} ({list(GROUP_CONFIGS.keys())})", flush=True)
print(f"USER_MEN_IDS (Managers): {USER_MEN_IDS}", flush=True)
print(f"USER_MAR_IDS (Marketing): {USER_MAR_IDS}", flush=True)
print(f"USER_ADMIN_ID: {USER_ADMIN_ID}", flush=True)
print(f"DB_HOST: {'SET' if DB_HOST else 'NOT SET'}", flush=True)
print(f"DB_NAME: {'SET' if DB_NAME else 'NOT SET'}", flush=True)
print(f"DB_USER: {'SET' if DB_USER else 'NOT SET'}", flush=True)
print(f"PORT: {PORT}", flush=True)
print("=" * 50, flush=True)

# -----------------------------------------------------------------------------
# Load Data Files
# -----------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def load_json_file(filename: str) -> Dict:
    """Load JSON file from disk."""
    filepath = os.path.join(BASE_DIR, filename)
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        print(f"Loaded {filename} successfully", flush=True)
        return data
    except Exception as e:
        print(f"Error loading {filename}: {e}", flush=True)
        return {}

# Load tests per course (tests/test_1.json, tests/test_2.json, ...)
TESTS_ALL_DATA: Dict[int, Dict] = {}
_tests_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tests")
for _ci in range(1, 5):
    _tf = os.path.join(_tests_dir, f"test_{_ci}.json")
    if os.path.isfile(_tf):
        try:
            with open(_tf, 'r', encoding='utf-8') as _f:
                TESTS_ALL_DATA[_ci] = json.load(_f)
            print(f"Loaded tests/test_{_ci}.json successfully", flush=True)
        except Exception as _e:
            print(f"Error loading tests/test_{_ci}.json: {_e}", flush=True)
    else:
        print(f"Warning: tests/test_{_ci}.json not found, course {_ci} tests disabled", flush=True)

TEXTS_DATA = load_json_file("texts.json")
FORM_DATA = load_json_file("form.json")
# Load final forms per course
FINAL_FORMS_DATA: Dict[int, Dict] = {}
for _course_num in [1, 2, 3]:
    _form_file = f"final_form_{_course_num}.json"
    _form_path = os.path.join(BASE_DIR, _form_file)
    if os.path.exists(_form_path):
        try:
            with open(_form_path, 'r', encoding='utf-8') as _f:
                FINAL_FORMS_DATA[_course_num] = json.load(_f)
            print(f"Loaded {_form_file} successfully", flush=True)
        except Exception as _e:
            print(f"Error loading {_form_file}: {_e}", flush=True)
    else:
        print(f"Warning: {_form_file} not found, course {_course_num} will use fallback", flush=True)

# Default/fallback: use course 1 form
FINAL_FORM_DATA = FINAL_FORMS_DATA.get(1, {})
CERTIFICATE_CONFIG = load_json_file("diploma_config.json")

# -----------------------------------------------------------------------------
# User Sessions Storage (in-memory)
# Format: {user_id: {"variant": N, "question": N, "score": N, "shuffled_answers": [...]}}
# -----------------------------------------------------------------------------
USER_SESSIONS: Dict[int, Dict] = {}

# -----------------------------------------------------------------------------
# Form Sessions Storage (in-memory)
# Format: {user_id: {"question": N, "answers": [str, str, ...]}}
# -----------------------------------------------------------------------------
FORM_SESSIONS: Dict[int, Dict] = {}

# -----------------------------------------------------------------------------
# Admin Search Sessions (in-memory)
# Format: {admin_id: {"step": str, "search_text": str, "results": [Dict], "page": int, "selected_user_id": int, "mode": str}}
# Steps: "search", "select_user", "select_course", "select_spins", "select_form"
# Modes: "access_survey", "fortune_wheel", "view_answers"
# -----------------------------------------------------------------------------
ADMIN_SEARCH_SESSIONS: Dict[int, Dict] = {}

# -----------------------------------------------------------------------------
# Final Form Sessions (in-memory)
# Format: {user_id: {"course": int, "step": str, "answers": Dict, "current_question": int/str}}
# Steps: "question", "check_data", "update_name", "update_case", "finished"
# -----------------------------------------------------------------------------
FINAL_FORM_SESSIONS: Dict[int, Dict] = {}

# -----------------------------------------------------------------------------
# Fortune Wheel Sessions (in-memory)
# Format: {user_id: {"course": int, "prize": str, "spin_available": bool}}
# -----------------------------------------------------------------------------
FORTUNE_WHEEL_SESSIONS: Dict[int, Dict] = {}

# -----------------------------------------------------------------------------
# Database Helper
# -----------------------------------------------------------------------------
class Database:
    """PostgreSQL database helper for user data."""
    
    def __init__(self):
        self.pool: Optional[asyncpg.Pool] = None
    
    async def init(self):
        """Initialize database connection pool and create table if not exists."""
        try:
            # Build connection string from Amvera variables
            # Format: postgresql://user:password@host:5432/database
            if DB_HOST and DB_NAME and DB_USER and DB_PASSWORD:
                conn_string = f"postgresql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:5432/{DB_NAME}"
            else:
                print("No database configuration found!", flush=True)
                print(f"DB_HOST: {'SET' if DB_HOST else 'NOT SET'}", flush=True)
                print(f"DB_NAME: {'SET' if DB_NAME else 'NOT SET'}", flush=True)
                print(f"DB_USER: {'SET' if DB_USER else 'NOT SET'}", flush=True)
                print(f"DB_PASSWORD: {'SET' if DB_PASSWORD else 'NOT SET'}", flush=True)
                return False
            
            print(f"Connecting to database {DB_NAME} at {DB_HOST}...", flush=True)
            self.pool = await asyncpg.create_pool(conn_string, min_size=2, max_size=10)
            
            # Create table if not exists
            async with self.pool.acquire() as conn:
                await conn.execute('''
                    CREATE TABLE IF NOT EXISTS users (
                        user_id BIGINT PRIMARY KEY,
                        user_name TEXT,
                        user_name_case TEXT DEFAULT '',
                        "komu_vydan" TEXT DEFAULT 'Иванову Ивану',
                        form_first INTEGER DEFAULT 0,
                        form_first_answer TEXT DEFAULT '',
                        test_book_1 INTEGER DEFAULT 0,
                        test_book_2 INTEGER DEFAULT 0,
                        test_book_3 INTEGER DEFAULT 0,
                        test_book_4 INTEGER DEFAULT 0,
                        practice_1 INTEGER DEFAULT 0,
                        practice_2 INTEGER DEFAULT 0,
                        practice_3 INTEGER DEFAULT 0,
                        practice_4 INTEGER DEFAULT 0,
                        access_survey_1 INTEGER DEFAULT 0,
                        access_survey_2 INTEGER DEFAULT 0,
                        access_survey_3 INTEGER DEFAULT 0,
                        access_survey_4 INTEGER DEFAULT 0,
                        form_end_1 TEXT DEFAULT '',
                        form_end_2 TEXT DEFAULT '',
                        form_end_3 TEXT DEFAULT '',
                        form_end_4 TEXT DEFAULT '',
                        diploma_1 INTEGER DEFAULT 0,
                        diploma_2 INTEGER DEFAULT 0,
                        diploma_3 INTEGER DEFAULT 0,
                        diploma_4 INTEGER DEFAULT 0,
                        fortune_wheel INTEGER DEFAULT 0,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                ''')
                print("Database table verified/created", flush=True)
            
            # Add new columns if they don't exist (for existing tables)
            async with self.pool.acquire() as conn:
                new_columns = [
                    ("user_name_case", "TEXT DEFAULT ''"),
                    ("\"komu_vydan\"", "TEXT DEFAULT 'Иванову Ивану'"),
                    ("form_first", "INTEGER DEFAULT 0"),
                    ("form_first_answer", "TEXT DEFAULT ''"),
                    ("test_book_1", "INTEGER DEFAULT 0"),
                    ("test_book_2", "INTEGER DEFAULT 0"),
                    ("test_book_3", "INTEGER DEFAULT 0"),
                    ("test_book_4", "INTEGER DEFAULT 0"),
                    ("practice_1", "INTEGER DEFAULT 0"),
                    ("practice_2", "INTEGER DEFAULT 0"),
                    ("practice_3", "INTEGER DEFAULT 0"),
                    ("practice_4", "INTEGER DEFAULT 0"),
                    ("access_survey_1", "INTEGER DEFAULT 0"),
                    ("access_survey_2", "INTEGER DEFAULT 0"),
                    ("access_survey_3", "INTEGER DEFAULT 0"),
                    ("access_survey_4", "INTEGER DEFAULT 0"),
                    ("form_end_1", "TEXT DEFAULT ''"),
                    ("form_end_2", "TEXT DEFAULT ''"),
                    ("form_end_3", "TEXT DEFAULT ''"),
                    ("form_end_4", "TEXT DEFAULT ''"),
                    ("diploma_1", "INTEGER DEFAULT 0"),
                    ("diploma_2", "INTEGER DEFAULT 0"),
                    ("diploma_3", "INTEGER DEFAULT 0"),
                    ("diploma_4", "INTEGER DEFAULT 0"),
                    ("fortune_wheel", "INTEGER DEFAULT 0"),
                ]
                for col_name, col_type in new_columns:
                    try:
                        await conn.execute(f"ALTER TABLE users ADD COLUMN IF NOT EXISTS {col_name} {col_type}")
                    except Exception as e:
                        print(f"Column {col_name} might already exist: {e}", flush=True)
                
                # Migrate BOOLEAN to INTEGER if needed with proper state transitions
                # Logic: 0 = inaccessible, 1 = accessible (show button), 2 = completed
                # Transition logic:
                #   - form_first = 2 → test_book_X = 1
                #   - test_book_X = 2 → practice_X = 1
                #   - access_survey_X = 1 (opened by manager), = 2 (completed)
                
                # Step 0: Alter column types from BOOLEAN to INTEGER
                boolean_columns = [
                    ("form_first", 1),  # default value for new users
                    ("test_book_1", 0),
                    ("test_book_2", 0),
                    ("test_book_3", 0),
                    ("test_book_4", 0),
                    ("practice_1", 0),
                    ("practice_2", 0),
                    ("practice_3", 0),
                    ("practice_4", 0),
                    ("access_survey_1", 0),
                    ("access_survey_2", 0),
                    ("access_survey_3", 0),
                    ("access_survey_4", 0),
                ]
                
                for col_name, default_val in boolean_columns:
                    try:
                        # Check if column is boolean type
                        col_type = await conn.fetchval(
                            "SELECT data_type FROM information_schema.columns WHERE table_name = 'users' AND column_name = $1",
                            col_name
                        )
                        if col_type == 'boolean':
                            # Step 1: Drop default
                            await conn.execute(f"ALTER TABLE users ALTER COLUMN {col_name} DROP DEFAULT")
                            # Step 2: Alter column type: TRUE -> 2, FALSE -> 0, NULL -> default
                            await conn.execute(f'''
                                ALTER TABLE users ALTER COLUMN {col_name} TYPE INTEGER 
                                USING CASE WHEN {col_name} = TRUE THEN 2 
                                           WHEN {col_name} = FALSE THEN 0 
                                           ELSE {default_val} END
                            ''')
                            # Step 3: Set new default
                            await conn.execute(f"ALTER TABLE users ALTER COLUMN {col_name} SET DEFAULT {default_val}")
                            print(f"Migration: {col_name} type changed from boolean to integer", flush=True)
                    except Exception as e:
                        print(f"Migration alter {col_name}: {e}", flush=True)
                
                # Step 1: Apply state transition logic (after type conversion)
                try:
                    # If form_first = 2 (completed), set test_book_X = 1 (accessible)
                    # But only if test_book_X is not already 2 (completed)
                    for i in range(1, 5):
                        await conn.execute(f'''UPDATE users SET test_book_{i} = 1 WHERE form_first = 2 AND (test_book_{i} = 0 OR test_book_{i} IS NULL)''')
                    print("Migration: test_book_X transition applied", flush=True)
                    
                    # If test_book_X = 2 (passed), set practice_X = 1 (accessible)
                    # But only if practice_X is not already 2 (completed)
                    for i in range(1, 5):
                        await conn.execute(f'''UPDATE users SET practice_{i} = 1 WHERE test_book_{i} = 2 AND (practice_{i} = 0 OR practice_{i} IS NULL)''')
                    print("Migration: practice_X transition applied", flush=True)
                    
                    print("Migration: State transitions completed", flush=True)
                except Exception as e:
                    print(f"Migration state transitions: {e}", flush=True)
                
                print("Database columns verified/added", flush=True)
            
            print("Database connection established", flush=True)
            return True
            
        except Exception as e:
            print(f"Database initialization error: {e}", flush=True)
            return False
    
    async def close(self):
        """Close database connection pool."""
        if self.pool:
            await self.pool.close()
            print("Database connection closed", flush=True)
    
    async def get_user(self, user_id: int) -> Optional[Dict]:
        """Get user by ID."""
        if not self.pool:
            return None
        
        try:
            async with self.pool.acquire() as conn:
                row = await conn.fetchrow(
                    "SELECT * FROM users WHERE user_id = $1", user_id
                )
                if row:
                    return dict(row)
                return None
        except Exception as e:
            print(f"Error getting user {user_id}: {e}", flush=True)
            return None
    
    async def get_all_users(self) -> List[Dict]:
        """Get all users from database."""
        if not self.pool:
            return []
        
        try:
            async with self.pool.acquire() as conn:
                rows = await conn.fetch("SELECT * FROM users ORDER BY created_at DESC")
                return [dict(row) for row in rows]
        except Exception as e:
            print(f"Error getting all users: {e}", flush=True)
            return []
    
    async def create_user(self, user_id: int, user_name: str) -> bool:
        """Create new user. By default, only form_first = 1 (accessible)."""
        if not self.pool:
            return False
        
        try:
            async with self.pool.acquire() as conn:
                await conn.execute('''
                    INSERT INTO users (user_id, user_name, "komu_vydan", form_first, form_first_answer, 
                                       test_book_1, test_book_2, test_book_3, test_book_4,
                                       practice_1, practice_2, practice_3, practice_4,
                                       access_survey_1, access_survey_2, access_survey_3, access_survey_4,
                                       diploma_1, diploma_2, diploma_3, diploma_4)
                    VALUES ($1, $2, 'Иванову Ивану', 1, '', 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
                    ON CONFLICT (user_id) DO UPDATE SET user_name = $2, updated_at = CURRENT_TIMESTAMP
                ''', user_id, user_name)
                print(f"User {user_id} created/updated in database", flush=True)
                return True
        except Exception as e:
            print(f"Error creating user {user_id}: {e}", flush=True)
            return False
    
    async def update_user_field(self, user_id: int, field: str, value: Any) -> bool:
        """Update a specific field for user."""
        if not self.pool:
            return False
        
        allowed_fields = [
            'form_first', 'form_first_answer', 
            'test_book_1', 'test_book_2', 'test_book_3', 'test_book_4', 
            'user_name', 'user_name_case', 'komu_vydan',
            'practice_1', 'practice_2', 'practice_3', 'practice_4',
            'access_survey_1', 'access_survey_2', 'access_survey_3', 'access_survey_4',
            'form_end_1', 'form_end_2', 'form_end_3', 'form_end_4',
            'diploma_1', 'diploma_2', 'diploma_3', 'diploma_4',
            'fortune_wheel'
        ]
        if field not in allowed_fields:
            return False
        
        try:
            async with self.pool.acquire() as conn:
                await conn.execute(
                    f'UPDATE users SET "{field}" = $1, updated_at = CURRENT_TIMESTAMP WHERE user_id = $2',
                    value, user_id
                )
                return True
        except Exception as e:
            print(f"Error updating user {user_id} field {field}: {e}", flush=True)
            return False
    
    async def increment_fortune_wheel(self, user_id: int, amount: int = 1) -> bool:
        """Increment fortune_wheel counter by amount (can be negative)."""
        if not self.pool:
            return False
        
        try:
            async with self.pool.acquire() as conn:
                await conn.execute(
                    "UPDATE users SET fortune_wheel = GREATEST(0, fortune_wheel + $1), updated_at = CURRENT_TIMESTAMP WHERE user_id = $2",
                    amount, user_id
                )
                return True
        except Exception as e:
            print(f"Error incrementing fortune_wheel for user {user_id}: {e}", flush=True)
            return False
    
    async def search_users_by_name(self, search_text: str) -> List[Dict]:
        """Search users by name (case-insensitive partial match).
        
        Note: PostgreSQL LOWER() doesn't work correctly with Cyrillic characters
        on some systems due to locale settings. We fetch all users and filter in Python.
        """
        if not self.pool:
            return []
        
        try:
            async with self.pool.acquire() as conn:
                rows = await conn.fetch(
                    "SELECT user_id, user_name FROM users ORDER BY user_name"
                )
                
                search_lower = search_text.lower()
                results = []
                for row in rows:
                    user_name = row['user_name']
                    if user_name and search_lower in user_name.lower():
                        results.append(dict(row))
                
                return results
        except Exception as e:
            logger.error(f"Error searching users by name '{search_text}': {e}")
            return []
    
    async def import_users(self, users: List[Dict]) -> Dict:
        """Import/update users from list. Returns stats: {created, updated, errors}."""
        if not self.pool:
            return {"created": 0, "updated": 0, "errors": 1, "message": "No database connection"}
        
        stats = {"created": 0, "updated": 0, "errors": 0}
        
        async with self.pool.acquire() as conn:
            for user in users:
                try:
                    user_id = user.get("user_id")
                    if not user_id:
                        stats["errors"] += 1
                        continue
                    
                    # Check if user exists
                    existing = await conn.fetchval(
                        "SELECT user_id FROM users WHERE user_id = $1", user_id
                    )
                    
                    if existing:
                        # Update existing user
                        await conn.execute('''
                            UPDATE users SET
                                user_name = $2,
                                "komu_vydan" = $3,
                                form_first = $4,
                                form_first_answer = $5,
                                fortune_wheel = $6,
                                test_book_1 = $7, test_book_2 = $8, test_book_3 = $9, test_book_4 = $10,
                                practice_1 = $11, practice_2 = $12, practice_3 = $13, practice_4 = $14,
                                access_survey_1 = $15, access_survey_2 = $16, access_survey_3 = $17, access_survey_4 = $18,
                                form_end_1 = $19, form_end_2 = $20, form_end_3 = $21, form_end_4 = $22,
                                diploma_1 = $23, diploma_2 = $24, diploma_3 = $25, diploma_4 = $26,
                                updated_at = CURRENT_TIMESTAMP
                            WHERE user_id = $1
                        ''',
                            user_id,
                            user.get("user_name", ""),
                            user.get("komu_vydan", ""),
                            user.get("form_first", 0),
                            user.get("form_first_answer", ""),
                            user.get("fortune_wheel", 0),
                            user.get("test_book_1", 0), user.get("test_book_2", 0),
                            user.get("test_book_3", 0), user.get("test_book_4", 0),
                            user.get("practice_1", 0), user.get("practice_2", 0),
                            user.get("practice_3", 0), user.get("practice_4", 0),
                            user.get("access_survey_1", 0), user.get("access_survey_2", 0),
                            user.get("access_survey_3", 0), user.get("access_survey_4", 0),
                            user.get("form_end_1", ""), user.get("form_end_2", ""),
                            user.get("form_end_3", ""), user.get("form_end_4", ""),
                            user.get("diploma_1", 0), user.get("diploma_2", 0),
                            user.get("diploma_3", 0), user.get("diploma_4", 0)
                        )
                        stats["updated"] += 1
                    else:
                        # Create new user
                        await conn.execute('''
                            INSERT INTO users (
                                user_id, user_name, "komu_vydan",
                                form_first, form_first_answer, fortune_wheel,
                                test_book_1, test_book_2, test_book_3, test_book_4,
                                practice_1, practice_2, practice_3, practice_4,
                                access_survey_1, access_survey_2, access_survey_3, access_survey_4,
                                form_end_1, form_end_2, form_end_3, form_end_4,
                                diploma_1, diploma_2, diploma_3, diploma_4
                            ) VALUES (
                                $1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14, $15, $16, $17, $18, $19, $20, $21, $22, $23, $24, $25, $26
                            )
                        ''',
                            user_id,
                            user.get("user_name", ""),
                            user.get("komu_vydan", ""),
                            user.get("form_first", 0),
                            user.get("form_first_answer", ""),
                            user.get("fortune_wheel", 0),
                            user.get("test_book_1", 0), user.get("test_book_2", 0),
                            user.get("test_book_3", 0), user.get("test_book_4", 0),
                            user.get("practice_1", 0), user.get("practice_2", 0),
                            user.get("practice_3", 0), user.get("practice_4", 0),
                            user.get("access_survey_1", 0), user.get("access_survey_2", 0),
                            user.get("access_survey_3", 0), user.get("access_survey_4", 0),
                            user.get("form_end_1", ""), user.get("form_end_2", ""),
                            user.get("form_end_3", ""), user.get("form_end_4", ""),
                            user.get("diploma_1", 0), user.get("diploma_2", 0),
                            user.get("diploma_3", 0), user.get("diploma_4", 0)
                        )
                        stats["created"] += 1
                        
                except Exception as e:
                    print(f"Error importing user {user.get('user_id')}: {e}", flush=True)
                    stats["errors"] += 1
        
        return stats

# Global database instance
db = Database()

# -----------------------------------------------------------------------------
# VK API Helper
# -----------------------------------------------------------------------------
class VKAPI:
    API_URL = "https://api.vk.com/method/"
    API_VERSION = "5.199"
    
    def __init__(self, token: str):
        self.token = token
        self.session: Optional[aiohttp.ClientSession] = None
    
    async def init(self):
        if self.session is None:
            self.session = aiohttp.ClientSession()
    
    async def close(self):
        if self.session:
            await self.session.close()
            self.session = None
    
    async def call(self, method: str, params: Dict[str, Any]) -> Dict[str, Any]:
        if not self.session:
            await self.init()
        
        params["access_token"] = self.token
        params["v"] = self.API_VERSION
        url = f"{self.API_URL}{method}"
        
        try:
            async with self.session.post(url, data=params) as resp:
                result = await resp.json()
                if "error" in result:
                    logger.error(f"VK API error: {result['error']}")
                return result
        except Exception as e:
            logger.exception(f"VK API call error: {e}")
            return {"error": str(e)}
    
    async def send_message(
        self, 
        user_id: int, 
        message: str, 
        keyboard: Optional[Dict] = None,
        peer_id: Optional[int] = None,
        attachment: Optional[str] = None
    ) -> Dict[str, Any]:
        params = {
            "message": message,
            "random_id": random.randint(0, 2**31 - 1)
        }
        
        if peer_id:
            params["peer_id"] = peer_id
        else:
            params["user_id"] = user_id
        
        if keyboard:
            params["keyboard"] = json.dumps(keyboard, ensure_ascii=False)
        
        if attachment:
            params["attachment"] = attachment
        
        return await self.call("messages.send", params)
    
    async def get_user_info(self, user_id: int) -> Dict[str, Any]:
        """Get user info by ID."""
        params = {"user_ids": user_id}
        return await self.call("users.get", params)
    
    async def get_conversations(self, offset: int = 0, count: int = 200) -> Dict[str, Any]:
        """Get conversations list with pagination."""
        params = {
            "offset": offset,
            "count": count,
            "filter": "all"
        }
        return await self.call("messages.getConversations", params)
    
    async def get_all_conversations(self) -> List[int]:
        """Get all user IDs from conversations with pagination."""
        all_user_ids = []
        offset = 0
        count = 200  # Max per request
        
        while True:
            result = await self.get_conversations(offset=offset, count=count)
            
            if "error" in result:
                print(f"Error getting conversations: {result['error']}", flush=True)
                break
            
            if "response" not in result:
                print(f"No response in conversations result: {result}", flush=True)
                break
            
            items = result["response"].get("items", [])
            if not items:
                break
            
            for item in items:
                conversation = item.get("conversation", {})
                peer = conversation.get("peer", {})
                peer_type = peer.get("type", "")
                peer_id = peer.get("id", 0)
                
                # Only user conversations (not groups, not chats)
                if peer_type == "user" and peer_id > 0:
                    all_user_ids.append(peer_id)
            
            total_count = result["response"].get("count", 0)
            offset += count
            
            print(f"Fetched {len(all_user_ids)} conversations, total available: {total_count}", flush=True)
            
            # Check if we got all
            if offset >= total_count:
                break
        
        return all_user_ids
    
    async def get_upload_server(self, peer_id: int) -> Optional[str]:
        """Get upload server URL for documents."""
        result = await self.call("docs.getMessagesUploadServer", {"peer_id": peer_id, "type": "doc"})
        if "response" in result and "upload_url" in result["response"]:
            return result["response"]["upload_url"]
        print(f"Error getting upload server: {result}", flush=True)
        return None
    
    async def upload_document(self, upload_url: str, file_data: bytes, filename: str) -> Optional[Dict]:
        """Upload document to VK."""
        if not self.session:
            await self.init()
        
        try:
            # Determine content type by file extension
            ext = filename.lower().split('.')[-1] if '.' in filename else ''
            content_types = {
                'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'png': 'image/png',
                'jpg': 'image/jpeg',
                'jpeg': 'image/jpeg',
                'pdf': 'application/pdf'
            }
            content_type = content_types.get(ext, 'application/octet-stream')
            
            form = aiohttp.FormData()
            form.add_field('file', file_data, filename=filename, content_type=content_type)
            
            async with self.session.post(upload_url, data=form) as resp:
                result = await resp.json()
                return result
        except Exception as e:
            print(f"Error uploading document: {e}", flush=True)
            return None
    
    async def save_document(self, file: str, title: str) -> Optional[Dict]:
        """Save uploaded document."""
        result = await self.call("docs.save", {"file": file, "title": title})
        return result
    
    async def send_document(self, peer_id: int, file_data: bytes, filename: str, message: str = "") -> bool:
        """Upload and send document to user."""
        try:
            # Get upload server
            upload_url = await self.get_upload_server(peer_id)
            if not upload_url:
                return False
            
            # Upload file
            upload_result = await self.upload_document(upload_url, file_data, filename)
            if not upload_result or "file" not in upload_result:
                print(f"Upload failed: {upload_result}", flush=True)
                return False
            
            # Save document
            save_result = await self.save_document(upload_result["file"], filename)
            if not save_result or "response" not in save_result:
                print(f"Save failed: {save_result}", flush=True)
                return False
            
            # Get document attachment string
            doc = save_result["response"].get("doc", save_result["response"].get("docs", [{}])[0] if "docs" in save_result["response"] else {})
            if not doc:
                print(f"No doc in response: {save_result}", flush=True)
                return False
            
            owner_id = doc.get("owner_id", doc.get("doc", {}).get("owner_id", ""))
            doc_id = doc.get("id", doc.get("doc", {}).get("id", ""))
            attachment = f"doc{owner_id}_{doc_id}"
            
            # Send message with attachment
            await self.send_message(
                user_id=0,
                message=message,
                peer_id=peer_id,
                attachment=attachment
            )
            
            return True
            
        except Exception as e:
            print(f"Error sending document: {e}", flush=True)
            return False
    
    async def download_document(self, url: str) -> Optional[bytes]:
        """Download document from VK by URL."""
        if not self.session:
            await self.init()
        
        try:
            async with self.session.get(url) as resp:
                if resp.status == 200:
                    return await resp.read()
                else:
                    print(f"Failed to download document: status {resp.status}", flush=True)
                    return None
        except Exception as e:
            print(f"Error downloading document: {e}", flush=True)
            return None


# -----------------------------------------------------------------------------
# Keyboard Builders
# -----------------------------------------------------------------------------
def create_main_menu_keyboard() -> Dict:
    """Main keyboard with only Menu button. Always visible."""
    return {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {"type": "text", "label": "Меню"},
                    "color": "primary"
                }
            ]
        ]
    }


def create_dynamic_menu_keyboard(
    is_admin: bool = False,
    is_manager: bool = False,
    is_marketing: bool = False,
    form_first: int = 0,
    test_book: int = 0,
    practice: int = 0,
    access_survey: int = 0,
    certificate: int = 0,
    fortune_wheel: int = 0,
    course_index: int = 1
) -> Dict:
    """
    Dynamic menu keyboard based on user state.
    
    Button-antagonists (only ONE can be shown at a time, value=1):
    1. Приветственная анкета (form_first=1)
    2. Тестирование (test_book_{N}=1)
    3. Сдал(а) практику (practice_{N}=1)
    4. Финальная анкета (access_survey_{N}=1)
    5. 📄 Скачать сертификат (diploma_{N}=1)
    
    N is determined by course_index (which VK group the user is in).
    Fortune wheel is separate (shown if fortune_wheel > 0)
    Special panels: Admin, Manager, Marketing
    """
    buttons = [
        [
            {
                "action": {"type": "text", "label": "Меню"},
                "color": "primary"
            }
        ]
    ]
    
    # Determine which action button to show (only one at a time)
    # Priority: check in order of progression
    action_button = None
    
    if form_first == 1:
        action_button = {
            "action": {"type": "text", "label": "Приветственная анкета"},
            "color": "positive"
        }
    elif test_book == 1:
        action_button = {
            "action": {"type": "text", "label": "Тестирование"},
            "color": "positive"
        }
    elif practice == 1:
        action_button = {
            "action": {"type": "text", "label": "Сдал(а) практику"},
            "color": "positive"
        }
    elif access_survey == 1:
        action_button = {
            "action": {"type": "text", "label": "Финальная анкета"},
            "color": "positive"
        }
    elif certificate == 1:
        action_button = {
            "action": {"type": "text", "label": "📄 Скачать сертификат"},
            "color": "positive"
        }
    
    # Add action button if exists
    if action_button:
        buttons[0].append(action_button)
    
    # Add fortune wheel button if spins available (separate from antagonists)
    if fortune_wheel > 0:
        buttons.append([
            {
                "action": {"type": "text", "label": f"Колесо фортуны ({fortune_wheel})"},
                "color": "positive"
            }
        ])
    
    # Add special panel buttons (Manager and Marketing)
    if is_manager:
        buttons.append([
            {
                "action": {"type": "text", "label": "Менеджер"},
                "color": "primary"
            }
        ])
    
    if is_marketing:
        buttons.append([
            {
                "action": {"type": "text", "label": "Маркетинг"},
                "color": "primary"
            }
        ])
    
    # Add admin button (for super admin only)
    if is_admin:
        buttons.append([
            {
                "action": {"type": "text", "label": "АДМИН"},
                "color": "negative"
            }
        ])
    
    return {
        "one_time": False,
        "inline": False,
        "buttons": buttons
    }

def create_admin_keyboard() -> Dict:
    """Admin panel keyboard with submenu."""
    return {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {"type": "text", "label": "Скачать базу"},
                    "color": "positive"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Загрузить базу"},
                    "color": "primary"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Открыть доступ к финальной анкете"},
                    "color": "primary"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Добавить вращений колеса фортуны"},
                    "color": "primary"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Посмотреть ответы на анкеты"},
                    "color": "primary"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Меню"},
                    "color": "secondary"
                }
            ]
        ]
    }

def create_manager_keyboard() -> Dict:
    """Manager panel keyboard for USER_MEN."""
    return {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {"type": "text", "label": "Открыть доступ к финальной анкете"},
                    "color": "primary"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Изменить поле Кому выдан"},
                    "color": "primary"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Добавить вращений колеса фортуны"},
                    "color": "primary"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Меню"},
                    "color": "secondary"
                }
            ]
        ]
    }

def create_marketing_keyboard() -> Dict:
    """Marketing panel keyboard for USER_MAR."""
    return {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {"type": "text", "label": "Посмотреть ответы на анкеты"},
                    "color": "primary"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Меню"},
                    "color": "secondary"
                }
            ]
        ]
    }

def create_spins_selection_keyboard() -> Dict:
    """Keyboard for selecting number of fortune wheel spins (1-10)."""
    buttons = []
    # Row 1: 1-5
    buttons.append([
        {"action": {"type": "text", "label": str(i)}, "color": "primary"} for i in range(1, 6)
    ])
    # Row 2: 6-10
    buttons.append([
        {"action": {"type": "text", "label": str(i)}, "color": "primary"} for i in range(6, 11)
    ])
    # Row 3: Menu
    buttons.append([
        {"action": {"type": "text", "label": "Меню"}, "color": "secondary"}
    ])
    return {
        "one_time": False,
        "inline": False,
        "buttons": buttons
    }

def create_form_selection_keyboard(user_data: Dict) -> Dict:
    """Keyboard for selecting which form answers to view."""
    buttons = []
    
    # Check which forms have answers
    if user_data.get("form_first_answer"):
        buttons.append([{"action": {"type": "text", "label": "Приветственная анкета"}, "color": "primary"}])
    
    for i in range(1, 5):
        if user_data.get(f"form_end_{i}"):
            buttons.append([{"action": {"type": "text", "label": f"Финальная анкета Курс {i}"}, "color": "primary"}])
    
    buttons.append([{"action": {"type": "text", "label": "Меню"}, "color": "secondary"}])
    
    return {
        "one_time": False,
        "inline": False,
        "buttons": buttons
    }

def create_user_search_keyboard(users: List[Dict], page: int = 0, per_page: int = 6) -> Dict:
    """Create keyboard with found users (6 per page with pagination)."""
    start_idx = page * per_page
    end_idx = start_idx + per_page
    page_users = users[start_idx:end_idx]
    total_pages = (len(users) + per_page - 1) // per_page
    
    buttons = []
    
    # Add user buttons (2 per row, 3 rows = 6 users)
    for i in range(0, len(page_users), 2):
        row = []
        for j in range(2):
            if i + j < len(page_users):
                user = page_users[i + j]
                user_name = user.get("user_name", "Unknown")[:20]  # Limit button text
                row.append({
                    "action": {"type": "text", "label": f"👤{user_name}"},
                    "color": "primary"
                })
        if row:
            buttons.append(row)
    
    # Add navigation buttons
    nav_row = [
        {
            "action": {"type": "text", "label": "Меню"},
            "color": "secondary"
        }
    ]
    
    if page > 0:
        nav_row.append({
            "action": {"type": "text", "label": "◀️ Назад"},
            "color": "primary"
        })
    
    if page < total_pages - 1:
        nav_row.append({
            "action": {"type": "text", "label": "Далее ▶️"},
            "color": "primary"
        })
    
    nav_row.append({
        "action": {"type": "text", "label": "🔄 Заново"},
        "color": "primary"
    })
    
    buttons.append(nav_row)
    
    return {
        "one_time": False,
        "inline": False,
        "buttons": buttons
    }

def create_course_selection_keyboard() -> Dict:
    """Keyboard for selecting course number."""
    return {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {"type": "text", "label": "1"},
                    "color": "primary"
                },
                {
                    "action": {"type": "text", "label": "2"},
                    "color": "primary"
                },
                {
                    "action": {"type": "text", "label": "3"},
                    "color": "primary"
                },
                {
                    "action": {"type": "text", "label": "4"},
                    "color": "primary"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Меню"},
                    "color": "secondary"
                }
            ]
        ]
    }

def create_access_action_keyboard() -> Dict:
    """Keyboard for opening/closing access."""
    return {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {"type": "text", "label": "🔓 Открыть"},
                    "color": "positive"
                },
                {
                    "action": {"type": "text", "label": "🔒 Закрыть"},
                    "color": "negative"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Меню"},
                    "color": "secondary"
                }
            ]
        ]
    }

def create_answer_keyboard(shuffled_answers: List[Dict]) -> Dict:
    """Keyboard with Menu button and answer buttons 1, 2, 3."""
    return {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {"type": "text", "label": "Меню"},
                    "color": "primary"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "1"},
                    "color": "primary"
                },
                {
                    "action": {"type": "text", "label": "2"},
                    "color": "primary"
                },
                {
                    "action": {"type": "text", "label": "3"},
                    "color": "primary"
                }
            ]
        ]
    }

def create_retry_keyboard() -> Dict:
    """Keyboard with Menu and Retry buttons for failed test."""
    restart_text = TEXTS_DATA.get("restart_button", "Пройти заново")
    return {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {"type": "text", "label": "Меню"},
                    "color": "primary"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": restart_text},
                    "color": "positive"
                }
            ]
        ]
    }

def create_form_keyboard() -> Dict:
    """Keyboard with only Menu button for form questions."""
    return {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {"type": "text", "label": "Меню"},
                    "color": "secondary"
                }
            ]
        ]
    }


def create_yes_no_keyboard() -> Dict:
    """Keyboard with Да/Нет buttons for final form questions (swapped colors: Да=negative, Нет=positive)."""
    return {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {"type": "text", "label": "Да"},
                    "color": "negative"
                },
                {
                    "action": {"type": "text", "label": "Нет"},
                    "color": "positive"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Меню"},
                    "color": "secondary"
                }
            ]
        ]
    }


def create_yes_no_keyboard_original() -> Dict:
    """Keyboard with Да/Нет buttons for final form questions (original colors: Да=positive, Нет=negative)."""
    return {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {"type": "text", "label": "Да"},
                    "color": "positive"
                },
                {
                    "action": {"type": "text", "label": "Нет"},
                    "color": "negative"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Меню"},
                    "color": "secondary"
                }
            ]
        ]
    }


def create_yes_no_unknown_keyboard() -> Dict:
    """Keyboard with Да/Нет/Не знаю buttons for final form questions (swapped colors: Да=negative, Нет=positive)."""
    return {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {"type": "text", "label": "Да"},
                    "color": "negative"
                },
                {
                    "action": {"type": "text", "label": "Нет"},
                    "color": "positive"
                },
                {
                    "action": {"type": "text", "label": "Не знаю"},
                    "color": "primary"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Меню"},
                    "color": "secondary"
                }
            ]
        ]
    }


def create_agree_disagree_keyboard() -> Dict:
    """Keyboard with Не против/Против buttons for final form questions."""
    return {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {"type": "text", "label": "Не против"},
                    "color": "positive"
                },
                {
                    "action": {"type": "text", "label": "Против"},
                    "color": "negative"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Меню"},
                    "color": "secondary"
                }
            ]
        ]
    }


def create_rating_keyboard(min_val: int = 1, max_val: int = 10) -> Dict:
    """Keyboard with rating buttons 1-10 for final form questions."""
    buttons = []
    
    # Create rows of 5 buttons each
    row1 = []
    row2 = []
    for i in range(min_val, max_val + 1):
        btn = {
            "action": {"type": "text", "label": str(i)},
            "color": "primary"
        }
        if i <= 5:
            row1.append(btn)
        else:
            row2.append(btn)
    
    buttons.append(row1)
    buttons.append(row2)
    buttons.append([
        {
            "action": {"type": "text", "label": "Меню"},
            "color": "secondary"
        }
    ])
    
    return {
        "one_time": False,
        "inline": False,
        "buttons": buttons
    }


def create_check_data_keyboard() -> Dict:
    """Keyboard for checking user data in final form (verify_name)."""
    return {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {"type": "text", "label": "Верно"},
                    "color": "positive"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Изменить ФИ"},
                    "color": "primary"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Меню"},
                    "color": "secondary"
                }
            ]
        ]
    }


def create_fortune_wheel_keyboard() -> Dict:
    """Keyboard for fortune wheel - spin or later."""
    return {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {"type": "text", "label": "Крутить"},
                    "color": "positive"
                },
                {
                    "action": {"type": "text", "label": "В другой раз"},
                    "color": "secondary"
                }
            ]
        ]
    }


def create_spin_wheel_keyboard() -> Dict:
    """Keyboard for spinning the wheel."""
    return {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {"type": "text", "label": "Испытать удачу"},
                    "color": "positive"
                }
            ],
            [
                {
                    "action": {"type": "text", "label": "Меню"},
                    "color": "secondary"
                }
            ]
        ]
    }


def create_final_form_open_keyboard() -> Dict:
    """Keyboard for open questions in final form (text input)."""
    return {
        "one_time": False,
        "inline": False,
        "buttons": [
            [
                {
                    "action": {"type": "text", "label": "Меню"},
                    "color": "secondary"
                }
            ]
        ]
    }


# -----------------------------------------------------------------------------
# Test Logic
# -----------------------------------------------------------------------------
def get_tests_data(course_index: int) -> Dict:
    """Get TESTS_DATA for a specific course. Returns empty dict if not found."""
    if course_index in TESTS_ALL_DATA:
        return TESTS_ALL_DATA[course_index]
    print(f"WARNING: No tests loaded for course {course_index}, available: {list(TESTS_ALL_DATA.keys())}", flush=True)
    return {}

def shuffle_answers(answers: List[Dict]) -> List[Dict]:
    """Shuffle answers and return new list with button numbers."""
    shuffled = deepcopy(answers)
    random.shuffle(shuffled)
    return shuffled

def get_random_variant(course_index: int = 1) -> int:
    """Get random variant number for a specific course."""
    td = get_tests_data(course_index)
    variants = td.get("variants", [])
    if variants:
        return random.randint(0, len(variants) - 1)
    return 0

def get_question(course_index: int, variant_idx: int, question_idx: int) -> Optional[Dict]:
    """Get question from variant for a specific course."""
    td = get_tests_data(course_index)
    variants = td.get("variants", [])
    if variant_idx < len(variants):
        questions = variants[variant_idx].get("questions", [])
        if question_idx < len(questions):
            return questions[question_idx]
    return None

def format_question_message(question: Dict, question_num: int, total: int) -> str:
    """Format question text for sending."""
    return f"Вопрос {question_num}/{total}:\n\n{question['question']}"

def get_correct_answer_text(question: Dict) -> str:
    """Get text of correct answer."""
    for answer in question.get("answers", []):
        if answer.get("is_correct", False):
            return answer.get("text", "")
    return ""


# -----------------------------------------------------------------------------
# Final Form Logic
# -----------------------------------------------------------------------------
def get_final_form_question(question_id, form_data: Optional[Dict] = None) -> Optional[Dict]:
    """Get question from final form by ID (int or str)."""
    data = form_data or FINAL_FORM_DATA
    questions = data.get("questions", [])
    for q in questions:
        qid = q.get("id")
        if str(qid) == str(question_id):
            return q
    return None


def get_next_question_id(current_id, answer: str = None, form_data: Optional[Dict] = None) -> Optional[Any]:
    """Get next question ID based on current question and answer."""
    question = get_final_form_question(current_id, form_data)
    if not question:
        return None
    
    # Check if there are branches based on answer
    branches = question.get("branches", {})
    if branches and answer:
        # For rating questions, check if answer is max or not
        if question.get("type") == "rating":
            max_val = question.get("max", 10)
            if answer == str(max_val):
                branch = branches.get("max", {})
            else:
                branch = branches.get("not_max", {})
            return branch.get("next_question")
        
        # For button questions, check exact match
        branch = branches.get(answer, {})
        if branch:
            action = branch.get("action")
            if action == "finish_form":
                return "finish"
            return branch.get("next_question")
    
    # Default next question
    next_q = question.get("next_question")
    if next_q == "finish_form":
        return "finish"
    return next_q


def select_prize_by_probability(prizes: List[Dict]) -> str:
    """Select a prize based on probability weights."""
    total = sum(p.get("probability", 0) for p in prizes)
    rand = random.randint(1, total)
    cumulative = 0
    for prize in prizes:
        cumulative += prize.get("probability", 0)
        if rand <= cumulative:
            return prize.get("name", "Приз")
    return prizes[-1].get("name", "Приз")


def get_sorted_prizes_list() -> str:
    """Get sorted prizes list for display (by probability, then alphabetically)."""
    prizes = FINAL_FORM_DATA.get("fortune_wheel", {}).get("prizes", [])
    # Sort by probability (ascending), then by name (alphabetically)
    sorted_prizes = sorted(prizes, key=lambda p: (p.get("probability", 0), p.get("name", "")))
    lines = []
    for p in sorted_prizes:
        lines.append(f"🎁 {p.get('name')} - вероятность {p.get('probability')}%")
    return "\n".join(lines)


def is_physical_prize(prize_name: str) -> bool:
    """Check if prize is a physical item that needs to be picked up."""
    physical_prizes = FINAL_FORM_DATA.get("fortune_wheel", {}).get("physical_prizes", [])
    return prize_name in physical_prizes


# -----------------------------------------------------------------------------
# Certificate Generation
# -----------------------------------------------------------------------------
def generate_certificate(course: int, name: str, date_str: str) -> Optional[bytes]:
    """Generate certificate image with course number, name and date.
    
    Args:
        course: Course number (1-4)
        name: User name for certificate
        date_str: Date string to put on certificate
        
    Returns:
        PNG image as bytes or None if failed
    """
    if not CERTIFICATE_CONFIG:
        logger.error("Certificate config not loaded")
        return None
    
    # Single template for all courses
    template_path = CERTIFICATE_CONFIG.get("template", "sert/sert.jpg")
    template_full_path = os.path.join(BASE_DIR, template_path)
    if not os.path.exists(template_full_path):
        logger.error(f"Template not found: {template_full_path}")
        return None
    
    # Load font
    font_path = os.path.join(BASE_DIR, CERTIFICATE_CONFIG.get("font", "sert/RussoOne-Regular.ttf"))
    if not os.path.exists(font_path):
        logger.error(f"Font not found: {font_path}")
        return None
    
    # Global font size and colors from config
    font_size = CERTIFICATE_CONFIG.get("font_size", 88)
    stroke_width = CERTIFICATE_CONFIG.get("stroke_width", 0)
    text_color_cfg = CERTIFICATE_CONFIG.get("text_color", {"r": 255, "g": 255, "b": 255})
    text_rgb = (int(text_color_cfg.get("r", 255)), int(text_color_cfg.get("g", 255)), int(text_color_cfg.get("b", 255)))
    
    # Course number text: "1-й", "2-й", "3-й", "4-й"
    course_labels = {1: "1-й", 2: "2-й", 3: "3-й", 4: "4-й"}
    course_text = course_labels.get(course, f"{course}-й")
    
    try:
        # Open template image
        img = Image.open(template_full_path)
        draw = ImageDraw.Draw(img)
        
        # Get field configs (pixel-based coordinates)
        fields = CERTIFICATE_CONFIG.get("fields", {})
        
        # Prepare field texts
        field_texts = {}
        for field_name, field_config in fields.items():
            if field_name == "course":
                field_texts[field_name] = course_text
            elif field_name == "name":
                field_texts[field_name] = name.upper() if field_config.get("uppercase", False) else name
            elif field_name == "date":
                field_texts[field_name] = date_str
            else:
                continue
        
        # Draw each field
        for field_name, field_config in fields.items():
            if field_name not in field_texts:
                continue
            
            text = field_texts[field_name]
            x = field_config.get("x", 893)
            y = field_config.get("y", 867)
            alignment = field_config.get("alignment", "center")
            vertical_align = field_config.get("vertical_align", "bottom")
            
            # Per-field font size, fallback to global
            field_font_size = field_config.get("font_size", font_size)
            field_font = ImageFont.truetype(font_path, field_font_size)
            
            # Get text bbox
            bbox = draw.textbbox((0, 0), text, font=field_font, stroke_width=stroke_width)
            text_width = bbox[2] - bbox[0]
            text_height = bbox[3] - bbox[1]
            
            # Horizontal alignment
            if alignment == "center":
                draw_x = x - text_width / 2
            elif alignment == "right":
                draw_x = x - text_width
            else:
                draw_x = x
            
            # Vertical alignment
            if vertical_align == "bottom":
                draw_y = y - text_height
            elif vertical_align == "center":
                draw_y = y - text_height / 2
            else:
                draw_y = y
            
            # Adjust for font baseline
            draw_y -= bbox[1]
            
            # Draw text (no stroke since stroke_width=0)
            draw.text(
                (int(draw_x), int(draw_y)),
                text,
                font=field_font,
                fill=text_rgb,
                stroke_width=stroke_width
            )
        
        # Save to bytes
        output = io.BytesIO()
        img.save(output, format="PNG")
        img.close()
        
        return output.getvalue()
        
    except Exception as e:
        logger.error(f"Error generating certificate: {e}")
        return None


# -----------------------------------------------------------------------------
# Database Export Helper
# -----------------------------------------------------------------------------
def create_users_xlsx(users: List[Dict]) -> bytes:
    """Create XLSX file with users data grouped by courses."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    
    # Define headers - grouped by courses
    headers = [
        "ID пользователя",
        "Имя",
        "Кому выдан",
        "Начальная анкета",
        "Ответ нач. анкеты",
        "Колесо фортуны",
        # Курс 1
        "Тест 1",
        "Практика 1",
        "Доступ к анкете 1",
        "Ответ анкеты 1",
        "Сертификат 1",
        # Курс 2
        "Тест 2",
        "Практика 2",
        "Доступ к анкете 2",
        "Ответ анкеты 2",
        "Сертификат 2",
        # Курс 3
        "Тест 3",
        "Практика 3",
        "Доступ к анкете 3",
        "Ответ анкеты 3",
        "Сертификат 3",
        # Курс 4
        "Тест 4",
        "Практика 4",
        "Доступ к анкете 4",
        "Ответ анкеты 4",
        "Сертификат 4",
        # Даты
        "Дата создания",
        "Дата обновления"
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_num, user in enumerate(users, 2):
        # Convert INTEGER status to Russian text
        # 0 = Нет, 1 = Доступен, 2 = Выполнено
        def status_ru(val):
            try:
                v = int(val) if val is not None else 0
                if v == 0:
                    return "Нет"
                elif v == 1:
                    return "Доступен"
                elif v == 2:
                    return "Выполнено"
                else:
                    return str(v)
            except:
                return str(val) if val else "Нет"
        
        def format_datetime(dt):
            if dt:
                if isinstance(dt, str):
                    return dt[:19] if len(dt) > 19 else dt
                return dt.strftime("%Y-%m-%d %H:%M:%S")
            return ""
        
        row_data = [
            user.get("user_id", ""),
            user.get("user_name", ""),
            user.get("komu_vydan", ""),
            status_ru(user.get("form_first", 0)),
            user.get("form_first_answer", ""),
            user.get("fortune_wheel", 0),
            # Курс 1
            status_ru(user.get("test_book_1", 0)),
            status_ru(user.get("practice_1", 0)),
            status_ru(user.get("access_survey_1", 0)),
            user.get("form_end_1", ""),
            status_ru(user.get("diploma_1", 0)),
            # Курс 2
            status_ru(user.get("test_book_2", 0)),
            status_ru(user.get("practice_2", 0)),
            status_ru(user.get("access_survey_2", 0)),
            user.get("form_end_2", ""),
            status_ru(user.get("diploma_2", 0)),
            # Курс 3
            status_ru(user.get("test_book_3", 0)),
            status_ru(user.get("practice_3", 0)),
            status_ru(user.get("access_survey_3", 0)),
            user.get("form_end_3", ""),
            status_ru(user.get("diploma_3", 0)),
            # Курс 4
            status_ru(user.get("test_book_4", 0)),
            status_ru(user.get("practice_4", 0)),
            status_ru(user.get("access_survey_4", 0)),
            user.get("form_end_4", ""),
            status_ru(user.get("diploma_4", 0)),
            # Даты
            format_datetime(user.get("created_at", "")),
            format_datetime(user.get("updated_at", ""))
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    
    # Adjust column widths
    column_widths = [15, 25, 25, 12, 40, 12, 12, 12, 15, 40, 10, 12, 12, 15, 40, 10, 12, 12, 15, 40, 10, 12, 12, 15, 40, 10, 20, 20]
    for col, width in enumerate(column_widths, 1):
        if col <= 26:
            col_letter = chr(64 + col)
        elif col <= 52:
            col_letter = f"A{chr(64 + col - 26)}"
        else:
            col_letter = f"B{chr(64 + col - 52)}"
        ws.column_dimensions[col_letter].width = width
    
    # Freeze first row
    ws.freeze_panes = "A2"
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def parse_users_xlsx(file_data: bytes) -> List[Dict]:
    """Parse XLSX file and return list of user dicts for database update."""
    from openpyxl import load_workbook
    
    wb = load_workbook(io.BytesIO(file_data))
    ws = wb.active
    
    # Get headers from first row
    headers = [cell.value for cell in ws[1]]
    
    # Map headers to column indices
    header_map = {h: i for i, h in enumerate(headers) if h}
    
    # Parse status from Russian text to INTEGER
    def parse_status(val):
        if val is None:
            return 0
        val_str = str(val).strip()
        if val_str == "Выполнено" or val_str == "Да":
            return 2
        elif val_str == "Доступен":
            return 1
        elif val_str == "Нет":
            return 0
        else:
            try:
                return int(val_str)
            except:
                return 0
    
    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        
        user = {
            "user_id": row[header_map.get("ID пользователя", 0)],
            "user_name": row[header_map.get("Имя", 1)] or "",
            "komu_vydan": row[header_map.get("Кому выдан", 2)] or "",
            "form_first": parse_status(row[header_map.get("Начальная анкета", 3)]),
            "form_first_answer": row[header_map.get("Ответ нач. анкеты", 4)] or "",
            "fortune_wheel": int(row[header_map.get("Колесо фортуны", 5)] or 0),
            # Курс 1
            "test_book_1": parse_status(row[header_map.get("Тест 1", 6)]),
            "practice_1": parse_status(row[header_map.get("Практика 1", 7)]),
            "access_survey_1": parse_status(row[header_map.get("Доступ к анкете 1", 8)]),
            "form_end_1": row[header_map.get("Ответ анкеты 1", 9)] or "",
            "diploma_1": parse_status(row[header_map.get("Сертификат 1", 10)]),
            # Курс 2
            "test_book_2": parse_status(row[header_map.get("Тест 2", 11)]),
            "practice_2": parse_status(row[header_map.get("Практика 2", 12)]),
            "access_survey_2": parse_status(row[header_map.get("Доступ к анкете 2", 13)]),
            "form_end_2": row[header_map.get("Ответ анкеты 2", 14)] or "",
            "diploma_2": parse_status(row[header_map.get("Сертификат 2", 15)]),
            # Курс 3
            "test_book_3": parse_status(row[header_map.get("Тест 3", 16)]),
            "practice_3": parse_status(row[header_map.get("Практика 3", 17)]),
            "access_survey_3": parse_status(row[header_map.get("Доступ к анкете 3", 18)]),
            "form_end_3": row[header_map.get("Ответ анкеты 3", 19)] or "",
            "diploma_3": parse_status(row[header_map.get("Сертификат 3", 20)]),
            # Курс 4
            "test_book_4": parse_status(row[header_map.get("Тест 4", 21)]),
            "practice_4": parse_status(row[header_map.get("Практика 4", 22)]),
            "access_survey_4": parse_status(row[header_map.get("Доступ к анкете 4", 23)]),
            "form_end_4": row[header_map.get("Ответ анкеты 4", 24)] or "",
            "diploma_4": parse_status(row[header_map.get("Сертификат 4", 25)]),
        }
        users.append(user)
    
    return users


# -----------------------------------------------------------------------------
# Web Server
# -----------------------------------------------------------------------------
class WebServer:
    def __init__(self):
        self.app = web.Application()
        # Create VKAPI instance per group
        self.vk_apis: Dict[int, VKAPI] = {}
        for gid, cfg in GROUP_CONFIGS.items():
            self.vk_apis[gid] = VKAPI(cfg["token"])
        # Default vk_api for backward compatibility
        self.vk_api = self.vk_apis.get(DEFAULT_GROUP_ID) if DEFAULT_GROUP_ID else None
        # Track current group for course routing
        self.current_group_id: int = DEFAULT_GROUP_ID
        self.current_course_index: int = GROUP_COURSE_MAP.get(DEFAULT_GROUP_ID, 1) if DEFAULT_GROUP_ID else 1
        self._setup_routes()
        print(f"WebServer initialized with {len(self.vk_apis)} group(s)", flush=True)

    def _setup_routes(self) -> None:
        self.app.router.add_post("/", self.vk_webhook)
        self.app.router.add_post("/webhook", self.vk_webhook)
        self.app.router.add_get("/", self.health)
        self.app.router.add_get("/health", self.health)
        print("Routes setup complete", flush=True)

    def _get_course_index(self) -> int:
        """Get the course index for the current group."""
        return GROUP_COURSE_MAP.get(self.current_group_id, 1)

    def _get_user_course_fields(self, user_data: Optional[Dict], course_index: int) -> Dict[str, int]:
        """Extract course-specific fields from user data.
        
        Returns dict with keys: test_book, practice, access_survey, diploma
        """
        if not user_data:
            return {"test_book": 0, "practice": 0, "access_survey": 0, "diploma": 0}
        return {
            "test_book": user_data.get(f"test_book_{course_index}", 0),
            "practice": user_data.get(f"practice_{course_index}", 0),
            "access_survey": user_data.get(f"access_survey_{course_index}", 0),
            "diploma": user_data.get(f"diploma_{course_index}", 0),
        }

    async def health(self, request: web.Request) -> web.Response:
        print("Health check requested", flush=True)
        return web.json_response({
            "status": "ok",
            "groups_configured": list(GROUP_CONFIGS.keys()),
            "user_men_ids": USER_MEN_IDS,
            "user_mar_ids": USER_MAR_IDS,
            "user_admin_id": USER_ADMIN_ID,
            "tests_loaded": bool(TESTS_ALL_DATA),
            "tests_courses": list(TESTS_ALL_DATA.keys()),
            "texts_loaded": bool(TEXTS_DATA),
            "form_loaded": bool(FORM_DATA),
            "final_form_loaded": bool(FINAL_FORM_DATA),
            "database_connected": db.pool is not None
        })

    async def vk_webhook(self, request: web.Request) -> web.Response:
        try:
            body = await request.text()
            print(f"POST received: {body[:500]}", flush=True)
            
            try:
                data = json.loads(body)
            except json.JSONDecodeError as e:
                print(f"JSON parse error: {e}", flush=True)
                return web.Response(text="invalid json", status=400)
            
            event_type = data.get("type", "")
            group_id = data.get("group_id", 0)
            
            print(f"Event: {event_type}, group: {group_id}", flush=True)
            
            # Check if group is configured
            if group_id not in GROUP_CONFIGS:
                print(f"WARNING: Unknown group_id {group_id}, configured groups: {list(GROUP_CONFIGS.keys())}", flush=True)
                return web.Response(text="ok")
            
            if event_type == "confirmation":
                confirm_token = GROUP_CONFIGS[group_id]["confirmation_token"]
                print(f"CONFIRMATION REQUEST for group {group_id} - returning: {confirm_token}", flush=True)
                return web.Response(text=confirm_token)
            
            if event_type == "message_new":
                # Switch to the correct VKAPI for this group
                self.vk_api = self.vk_apis.get(group_id)
                if not self.vk_api:
                    print(f"ERROR: No VKAPI for group {group_id}", flush=True)
                    return web.Response(text="ok")
                # Track current group for course routing
                self.current_group_id = group_id
                self.current_course_index = GROUP_COURSE_MAP.get(group_id, 1)
                print(f"Handling message_new for group {group_id} (course {self.current_course_index})", flush=True)
                await self._handle_message_new(data)
                return web.Response(text="ok")
            
            print(f"Unknown event: {event_type}", flush=True)
            return web.Response(text="ok")
            
        except Exception as e:
            print(f"ERROR: {e}", flush=True)
            return web.Response(text="ok")

    async def _send_user_menu(self, user_id: int, peer_id: int, message: str = "Выберите действие:") -> None:
        """Send user's main menu with dynamic keyboard based on their state."""
        # Get user data
        user_data = await db.get_user(user_id)
        course_index = self._get_course_index()
        
        # Check user roles
        is_admin = (user_id == USER_ADMIN_ID)
        is_manager = (user_id in USER_MEN_IDS)
        is_marketing = (user_id in USER_MAR_IDS)
        
        # Get user state for current course
        form_first_status = user_data.get("form_first", 0) if user_data else 0
        cf = self._get_user_course_fields(user_data, course_index)
        fortune_wheel_spins = user_data.get("fortune_wheel", 0) if user_data else 0
        
        keyboard = create_dynamic_menu_keyboard(
            is_admin=is_admin,
            is_manager=is_manager,
            is_marketing=is_marketing,
            form_first=form_first_status,
            test_book=cf["test_book"],
            practice=cf["practice"],
            access_survey=cf["access_survey"],
            certificate=cf["diploma"],
            fortune_wheel=fortune_wheel_spins,
            course_index=course_index
        )
        
        await self.vk_api.send_message(
            user_id=user_id,
            message=message,
            peer_id=peer_id,
            keyboard=keyboard
        )

    async def _handle_message_new(self, data: Dict) -> None:
        try:
            if not self.vk_api:
                print("No VK API", flush=True)
                return
            
            await self.vk_api.init()
            
            obj = data.get("object", {})
            message = obj.get("message", obj)
            
            user_id = message.get("from_id", message.get("user_id", 0))
            peer_id = message.get("peer_id", user_id)
            text = message.get("text", "").strip()
            attachments = message.get("attachments", [])
            
            print(f"User {user_id}: {text}", flush=True)
            if attachments:
                print(f"Attachments: {len(attachments)}", flush=True)
            
            # Check user roles
            is_admin = (user_id == USER_ADMIN_ID)
            is_manager = (user_id in USER_MEN_IDS)
            is_marketing = (user_id in USER_MAR_IDS)
            
            # Handle file upload for admin (Excel import)
            if is_admin and attachments:
                # Check if admin is expecting file upload
                admin_session = ADMIN_SEARCH_SESSIONS.get(user_id, {})
                if admin_session.get("step") == "awaiting_file_upload":
                    # Find xlsx attachment
                    for att in attachments:
                        if att.get("type") == "doc":
                            doc = att.get("doc", {})
                            filename = doc.get("title", "")
                            ext = doc.get("ext", "")
                            url = doc.get("url", "")
                            
                            print(f"Document: {filename}, ext: {ext}", flush=True)
                            
                            if ext == "xlsx" or filename.endswith(".xlsx"):
                                # Download file
                                await self.vk_api.send_message(
                                    user_id=user_id,
                                    message="⏳ Загружаю и обрабатываю файл...",
                                    peer_id=peer_id
                                )
                                
                                file_data = await self.vk_api.download_document(url)
                                if file_data:
                                    await self._handle_import_db(user_id, peer_id, file_data, filename)
                                else:
                                    await self.vk_api.send_message(
                                        user_id=user_id,
                                        message="❌ Не удалось скачать файл.",
                                        peer_id=peer_id,
                                        keyboard=create_admin_keyboard()
                                    )
                                
                                # Clear admin session
                                if user_id in ADMIN_SEARCH_SESSIONS:
                                    del ADMIN_SEARCH_SESSIONS[user_id]
                                return
                    
                    # No xlsx found
                    await self.vk_api.send_message(
                        user_id=user_id,
                        message="❌ Файл не найден или неверный формат. Пожалуйста, отправьте файл .xlsx",
                        peer_id=peer_id,
                        keyboard=create_admin_keyboard()
                    )
                    return
            
            # Handle "Start" button
            # Always handle "Начать" — clear any active sessions and show menu
            if text.lower() in ["начать", "start", "/start"]:
                # Clear any existing sessions
                if user_id in USER_SESSIONS:
                    del USER_SESSIONS[user_id]
                if user_id in FORM_SESSIONS:
                    del FORM_SESSIONS[user_id]
                if user_id in ADMIN_SEARCH_SESSIONS:
                    del ADMIN_SEARCH_SESSIONS[user_id]
                if user_id in FINAL_FORM_SESSIONS:
                    del FINAL_FORM_SESSIONS[user_id]
                if user_id in FORTUNE_WHEEL_SESSIONS:
                    del FORTUNE_WHEEL_SESSIONS[user_id]
                
                # Get user name from VK
                user_name = f"ID{user_id}"
                try:
                    user_info = await self.vk_api.get_user_info(user_id)
                    if "response" in user_info and user_info["response"]:
                        first_name = user_info["response"][0].get("first_name", "")
                        last_name = user_info["response"][0].get("last_name", "")
                        user_name = f"{first_name} {last_name}"
                except Exception as e:
                    print(f"Error getting user info: {e}", flush=True)
                
                # Check/create user in database
                user_data = await db.get_user(user_id)
                if not user_data:
                    # Create new user (form_first = 1 by default)
                    await db.create_user(user_id, user_name)
                    user_data = await db.get_user(user_id)
                
                # Get status - now using INTEGER values
                # 0 = inaccessible, 1 = accessible (show button), 2 = completed
                course_index = self._get_course_index()
                form_first_status = user_data.get("form_first", 0) if user_data else 0
                cf = self._get_user_course_fields(user_data, course_index)
                fortune_wheel_spins = user_data.get("fortune_wheel", 0) if user_data else 0
                
                # Choose appropriate keyboard (same logic as Menu)
                # Use dynamic menu keyboard - shows one action button + fortune wheel if available
                keyboard = create_dynamic_menu_keyboard(
                    is_admin=is_admin,
                    is_manager=is_manager,
                    is_marketing=is_marketing,
                    form_first=form_first_status,
                    test_book=cf["test_book"],
                    practice=cf["practice"],
                    access_survey=cf["access_survey"],
                    certificate=cf["diploma"],
                    fortune_wheel=fortune_wheel_spins,
                    course_index=course_index
                )
                
                await self.vk_api.send_message(
                    user_id=user_id,
                    message=TEXTS_DATA.get("welcome_message", "Добро пожаловать! Используйте кнопки клавиатуры."),
                    peer_id=peer_id,
                    keyboard=keyboard
                )
                return
            
            # Handle "Меню" button - always available
            if text.lower() == "меню":
                # Clear test session if user is in the middle of a test
                if user_id in USER_SESSIONS:
                    del USER_SESSIONS[user_id]
                    print(f"Test session cleared for user {user_id} - returned to menu", flush=True)
                
                # Clear form session if user is in the middle of form
                if user_id in FORM_SESSIONS:
                    del FORM_SESSIONS[user_id]
                    print(f"Form session cleared for user {user_id} - returned to menu", flush=True)
                
                # Clear admin search session
                if user_id in ADMIN_SEARCH_SESSIONS:
                    del ADMIN_SEARCH_SESSIONS[user_id]
                    print(f"Admin search session cleared for user {user_id} - returned to menu", flush=True)
                
                # Clear final form session
                if user_id in FINAL_FORM_SESSIONS:
                    del FINAL_FORM_SESSIONS[user_id]
                    print(f"Final form session cleared for user {user_id} - returned to menu", flush=True)
                
                # Clear fortune wheel session
                if user_id in FORTUNE_WHEEL_SESSIONS:
                    del FORTUNE_WHEEL_SESSIONS[user_id]
                    print(f"Fortune wheel session cleared for user {user_id} - returned to menu", flush=True)
                
                # Check/create user by ID
                user_data = await db.get_user(user_id)
                if not user_data:
                    # Get user name from VK for new user
                    user_name = f"ID{user_id}"
                    try:
                        user_info = await self.vk_api.get_user_info(user_id)
                        if "response" in user_info and user_info["response"]:
                            first_name = user_info["response"][0].get("first_name", "")
                            last_name = user_info["response"][0].get("last_name", "")
                            user_name = f"{first_name} {last_name}"
                    except Exception as e:
                        print(f"Error getting user info: {e}", flush=True)
                    
                    # Create new user
                    await db.create_user(user_id, user_name)
                    print(f"New user {user_id} ({user_name}) created via Menu button", flush=True)
                    user_data = await db.get_user(user_id)
                
                # Get status - now using INTEGER values
                # 0 = inaccessible, 1 = accessible (show button), 2 = completed
                course_index = self._get_course_index()
                form_first_status = user_data.get("form_first", 0) if user_data else 0
                cf = self._get_user_course_fields(user_data, course_index)
                fortune_wheel_spins = user_data.get("fortune_wheel", 0) if user_data else 0
                
                # Choose appropriate keyboard based on user state
                # Use dynamic menu keyboard - shows one action button + fortune wheel if available
                keyboard = create_dynamic_menu_keyboard(
                    is_admin=is_admin,
                    is_manager=is_manager,
                    is_marketing=is_marketing,
                    form_first=form_first_status,
                    test_book=cf["test_book"],
                    practice=cf["practice"],
                    access_survey=cf["access_survey"],
                    certificate=cf["diploma"],
                    fortune_wheel=fortune_wheel_spins,
                    course_index=course_index
                )
                
                await self.vk_api.send_message(
                    user_id=user_id,
                    message=TEXTS_DATA.get("menu_select_action", "Выберите действие:"),
                    peer_id=peer_id,
                    keyboard=keyboard
                )
                return
            
            # Handle "АДМИН" button - only for USER_ADMIN (show admin menu)
            if text.lower() == "админ" and is_admin:
                await self.vk_api.send_message(
                    user_id=user_id,
                    message=TEXTS_DATA.get("admin_panel_title", "🔧 Админ-панель:\n\nВыберите действие:"),
                    peer_id=peer_id,
                    keyboard=create_admin_keyboard()
                )
                return
            
            # Handle "Менеджер" button - for USER_MEN (show manager menu)
            if text.lower() == "менеджер" and is_manager:
                await self.vk_api.send_message(
                    user_id=user_id,
                    message="📋 Панель менеджера:\n\nВыберите действие:",
                    peer_id=peer_id,
                    keyboard=create_manager_keyboard()
                )
                return
            
            # Handle "Маркетинг" button - for USER_MAR (show marketing menu)
            if text.lower() == "маркетинг" and is_marketing:
                await self.vk_api.send_message(
                    user_id=user_id,
                    message="📊 Панель маркетинга:\n\nВыберите действие:",
                    peer_id=peer_id,
                    keyboard=create_marketing_keyboard()
                )
                return
            
            # Handle "Скачать базу" button
            if text.lower() == "скачать базу" and is_admin:
                await self._handle_download_db(user_id, peer_id)
                return
            
            # Handle "Загрузить базу" button
            if text.lower() == "загрузить базу" and is_admin:
                await self.vk_api.send_message(
                    user_id=user_id,
                    message="📤 Для загрузки базы данных отправьте Excel файл (.xlsx) в этот чат.\n\nФайл должен иметь ту же структуру, что и скачанный файл базы.",
                    peer_id=peer_id,
                    keyboard=create_admin_keyboard()
                )
                # Set admin session to expect file upload
                ADMIN_SEARCH_SESSIONS[user_id] = {
                    "step": "awaiting_file_upload",
                    "search_text": "",
                    "results": [],
                    "page": 0,
                    "selected_user_id": None,
                    "selected_course": None
                }
                return
            
            # Handle "Открыть доступ к финальной анкете" button - for admin and managers
            if text.lower() == "открыть доступ к финальной анкете" and (is_admin or is_manager):
                # Clear any existing admin search session
                if user_id in ADMIN_SEARCH_SESSIONS:
                    del ADMIN_SEARCH_SESSIONS[user_id]
                
                await self.vk_api.send_message(
                    user_id=user_id,
                    message=TEXTS_DATA.get("admin_search_prompt", "🔍 Поиск пользователя:\n\nВведите имя или часть имени для поиска:"),
                    peer_id=peer_id,
                    keyboard=create_main_menu_keyboard()
                )
                # Start admin search session
                ADMIN_SEARCH_SESSIONS[user_id] = {
                    "step": "search",
                    "mode": "access_survey",
                    "search_text": "",
                    "results": [],
                    "page": 0,
                    "selected_user_id": None,
                    "selected_course": None
                }
                return
            
            # Handle "Добавить вращений колеса фортуны" button - for admin and managers
            if text.lower() == "добавить вращений колеса фортуны" and (is_admin or is_manager):
                # Clear any existing admin search session
                if user_id in ADMIN_SEARCH_SESSIONS:
                    del ADMIN_SEARCH_SESSIONS[user_id]
                
                await self.vk_api.send_message(
                    user_id=user_id,
                    message="🎡 Добавление вращений колеса фортуны:\n\n🔍 Введите имя или часть имени для поиска пользователя:",
                    peer_id=peer_id,
                    keyboard=create_main_menu_keyboard()
                )
                # Start admin search session for fortune wheel
                ADMIN_SEARCH_SESSIONS[user_id] = {
                    "step": "search",
                    "mode": "fortune_wheel",
                    "search_text": "",
                    "results": [],
                    "page": 0,
                    "selected_user_id": None,
                    "selected_course": None
                }
                return
            
            # Handle "Изменить поле Кому выдан" button - for admin and managers
            if text.lower() == "изменить поле кому выдан" and (is_admin or is_manager):
                # Clear any existing admin search session
                if user_id in ADMIN_SEARCH_SESSIONS:
                    del ADMIN_SEARCH_SESSIONS[user_id]
                
                await self.vk_api.send_message(
                    user_id=user_id,
                    message="✏️ Изменение поля \"Кому выдан\":\n\n🔍 Введите имя или часть имени для поиска пользователя:",
                    peer_id=peer_id,
                    keyboard=create_main_menu_keyboard()
                )
                # Start admin search session for editing komu_vydan
                ADMIN_SEARCH_SESSIONS[user_id] = {
                    "step": "search",
                    "mode": "edit_komu_vydan",
                    "search_text": "",
                    "results": [],
                    "page": 0,
                    "selected_user_id": None,
                    "selected_course": None
                }
                return
            
            # Handle "Посмотреть ответы на анкеты" button - for admin and marketing
            if text.lower() == "посмотреть ответы на анкеты" and (is_admin or is_marketing):
                # Clear any existing admin search session
                if user_id in ADMIN_SEARCH_SESSIONS:
                    del ADMIN_SEARCH_SESSIONS[user_id]
                
                await self.vk_api.send_message(
                    user_id=user_id,
                    message="📋 Просмотр ответов на анкеты:\n\n🔍 Введите имя или часть имени для поиска пользователя:",
                    peer_id=peer_id,
                    keyboard=create_main_menu_keyboard()
                )
                # Start admin search session for viewing answers
                ADMIN_SEARCH_SESSIONS[user_id] = {
                    "step": "search",
                    "mode": "view_answers",
                    "search_text": "",
                    "results": [],
                    "page": 0,
                    "selected_user_id": None,
                    "selected_course": None
                }
                return
            
            # Handle admin search session
            if (is_admin or is_manager or is_marketing) and user_id in ADMIN_SEARCH_SESSIONS:
                session = ADMIN_SEARCH_SESSIONS[user_id]
                
                # Handle pagination - "Далее"
                if text.startswith("Далее"):
                    session["page"] += 1
                    await self._show_search_results(user_id, peer_id)
                    return
                
                # Handle pagination - "Назад"
                if text.startswith("◀️") or text == "Назад":
                    session["page"] = max(0, session["page"] - 1)
                    await self._show_search_results(user_id, peer_id)
                    return
                
                # Handle "Заново" - restart search with same mode
                if text.startswith("🔄") or text == "Заново":
                    # Preserve the mode before clearing session
                    current_mode = session.get("mode", "access_survey")
                    del ADMIN_SEARCH_SESSIONS[user_id]
                    ADMIN_SEARCH_SESSIONS[user_id] = {
                        "step": "search",
                        "mode": current_mode,  # Keep the same mode
                        "search_text": "",
                        "results": [],
                        "page": 0,
                        "selected_user_id": None,
                        "selected_course": None
                    }
                    
                    # Show appropriate message based on mode
                    if current_mode == "fortune_wheel":
                        message = "🎡 Добавление вращений колеса фортуны:\n\n🔍 Введите имя или часть имени для поиска пользователя:"
                    elif current_mode == "edit_komu_vydan":
                        message = '✏️ Изменение поля "Кому выдан":\n\n🔍 Введите имя или часть имени для поиска пользователя:'
                    elif current_mode == "view_answers":
                        message = "📋 Просмотр ответов на анкеты:\n\n🔍 Введите имя или часть имени для поиска пользователя:"
                    else:
                        message = TEXTS_DATA.get("admin_search_prompt_retry", "🔍 Введите имя или часть имени для поиска:")
                    
                    await self.vk_api.send_message(
                        user_id=user_id,
                        message=message,
                        peer_id=peer_id,
                        keyboard=create_main_menu_keyboard()
                    )
                    return
                
                # Handle user selection (username button)
                if text.startswith("👤"):
                    # Extract user name from button and find user in results
                    selected_name = text[1:].strip()  # Remove 👤 prefix
                    # Find user in results by name
                    for user in session["results"]:
                        if user.get("user_name", "").startswith(selected_name):
                            session["selected_user_id"] = user.get("user_id")
                            break
                    
                    if session["selected_user_id"]:
                        mode = session.get("mode", "access_survey")
                        
                        if mode == "access_survey":
                            # Show course selection for access survey
                            session["step"] = "select_course"
                            msg_template = TEXTS_DATA.get("admin_select_course", "Выберите курс для пользователя:\n{user_name}")
                            await self.vk_api.send_message(
                                user_id=user_id,
                                message=msg_template.format(user_name=selected_name),
                                peer_id=peer_id,
                                keyboard=create_course_selection_keyboard()
                            )
                        
                        elif mode == "fortune_wheel":
                            # Show spins selection for fortune wheel
                            session["step"] = "select_spins"
                            await self.vk_api.send_message(
                                user_id=user_id,
                                message=f"🎡 Выберите количество вращений для пользователя:\n{selected_name}",
                                peer_id=peer_id,
                                keyboard=create_spins_selection_keyboard()
                            )
                        
                        elif mode == "view_answers":
                            # Show form selection for viewing answers
                            session["step"] = "select_form"
                            selected_user = await db.get_user(session["selected_user_id"])
                            
                            if selected_user:
                                # Check which forms have answers
                                has_forms = False
                                if selected_user.get("form_first_answer"):
                                    has_forms = True
                                for i in range(1, 5):
                                    if selected_user.get(f"form_end_{i}"):
                                        has_forms = True
                                
                                if has_forms:
                                    await self.vk_api.send_message(
                                        user_id=user_id,
                                        message=f"📋 Выберите анкету для просмотра:\n{selected_name}",
                                        peer_id=peer_id,
                                        keyboard=create_form_selection_keyboard(selected_user)
                                    )
                                else:
                                    await self.vk_api.send_message(
                                        user_id=user_id,
                                        message=f"❌ У пользователя {selected_name} нет заполненных анкет.",
                                        peer_id=peer_id,
                                        keyboard=create_main_menu_keyboard()
                                    )
                                    del ADMIN_SEARCH_SESSIONS[user_id]
                        
                        elif mode == "edit_komu_vydan":
                            # Ask for new komu_vydan value
                            selected_user = await db.get_user(session["selected_user_id"])
                            current_komu_vydan = selected_user.get("komu_vydan", "") if selected_user else ""
                            session["step"] = "awaiting_komu_vydan"
                            msg_template = TEXTS_DATA.get("edit_komu_vydan_prompt", "✏️ Текущее значение \"Кому выдан\" для пользователя {user_name}:\n\n{current_value}\n\nВведите новое значение:")
                            await self.vk_api.send_message(
                                user_id=user_id,
                                message=msg_template.format(user_name=selected_name, current_value=current_komu_vydan),
                                peer_id=peer_id,
                                keyboard=create_main_menu_keyboard()
                            )
                    return
                
                # Handle spins selection (1-10) for fortune wheel
                if session.get("mode") == "fortune_wheel" and session["step"] == "select_spins" and text.isdigit():
                    spins = int(text)
                    if 1 <= spins <= 10:
                        target_user_id = session["selected_user_id"]
                        
                        # Get current spins and add new ones
                        target_user = await db.get_user(target_user_id)
                        current_spins = target_user.get("fortune_wheel", 0) if target_user else 0
                        new_spins = current_spins + spins
                        
                        # Update database
                        await db.update_user_field(target_user_id, "fortune_wheel", new_spins)
                        
                        user_name = target_user.get("user_name", "Unknown") if target_user else "Unknown"
                        
                        # Determine which keyboard to show
                        keyboard = create_manager_keyboard() if is_manager else create_admin_keyboard()
                        
                        await self.vk_api.send_message(
                            user_id=user_id,
                            message=f"✅ Добавлено {spins} вращений колеса фортуны для:\n{user_name}\n\nТеперь всего: {new_spins}",
                            peer_id=peer_id,
                            keyboard=keyboard
                        )
                        
                        del ADMIN_SEARCH_SESSIONS[user_id]
                        return
                
                # Handle awaiting komu_vydan text input
                if session.get("mode") == "edit_komu_vydan" and session["step"] == "awaiting_komu_vydan" and text:
                    target_user_id = session["selected_user_id"]
                    
                    # Update komu_vydan in database
                    await db.update_user_field(target_user_id, "komu_vydan", text)
                    
                    target_user = await db.get_user(target_user_id)
                    user_name = target_user.get("user_name", "Unknown") if target_user else "Unknown"
                    
                    # Determine which keyboard to show
                    keyboard = create_manager_keyboard() if is_manager else create_admin_keyboard()
                    
                    msg_template = TEXTS_DATA.get("edit_komu_vydan_success", '✅ Поле "Кому выдан" обновлено для пользователя {user_name}:\n\n{new_value}')
                    await self.vk_api.send_message(
                        user_id=user_id,
                        message=msg_template.format(user_name=user_name, new_value=text),
                        peer_id=peer_id,
                        keyboard=keyboard
                    )
                    
                    del ADMIN_SEARCH_SESSIONS[user_id]
                    return
                
                # Handle form selection for viewing answers
                if session.get("mode") == "view_answers" and session["step"] == "select_form":
                    target_user_id = session["selected_user_id"]
                    target_user = await db.get_user(target_user_id)
                    user_name = target_user.get("user_name", "Unknown") if target_user else "Unknown"
                    
                    answers_text = ""
                    
                    if text == "Приветственная анкета":
                        answers_text = target_user.get("form_first_answer", "") if target_user else ""
                        if answers_text:
                            answers_text = f"📋 ПРИВЕТСТВЕННАЯ АНКЕТА\n👤 Пользователь: {user_name}\n\n{answers_text}"
                    
                    elif text.startswith("Финальная анкета Курс"):
                        course_num = text.split()[-1]
                        answers_text = target_user.get(f"form_end_{course_num}", "") if target_user else ""
                        if answers_text:
                            answers_text = f"📋 ФИНАЛЬНАЯ АНКЕТА КУРС {course_num}\n👤 Пользователь: {user_name}\n\n{answers_text}"
                    
                    if answers_text:
                        # Split message if too long (VK limit ~4096 chars)
                        if len(answers_text) > 4000:
                            chunks = [answers_text[i:i+4000] for i in range(0, len(answers_text), 4000)]
                            for chunk in chunks:
                                await self.vk_api.send_message(
                                    user_id=user_id,
                                    message=chunk,
                                    peer_id=peer_id,
                                    keyboard=create_main_menu_keyboard()
                                )
                        else:
                            await self.vk_api.send_message(
                                user_id=user_id,
                                message=answers_text,
                                peer_id=peer_id,
                                keyboard=create_main_menu_keyboard()
                            )
                    else:
                        await self.vk_api.send_message(
                            user_id=user_id,
                            message=f"❌ Ответы не найдены.",
                            peer_id=peer_id,
                            keyboard=create_main_menu_keyboard()
                        )
                    
                    del ADMIN_SEARCH_SESSIONS[user_id]
                    return
                
                # Handle course selection (1-4) when selecting user
                if session["step"] == "select_course" and text in ["1", "2", "3", "4"]:
                    session["selected_course"] = int(text)
                    session["step"] = "confirm_action"
                    
                    # Get selected user info
                    selected_user = await db.get_user(session["selected_user_id"])
                    user_name = selected_user.get("user_name", "Unknown") if selected_user else "Unknown"
                    
                    # Show access control options (no practice check needed)
                    msg_template = TEXTS_DATA.get("admin_access_action", "Выберите действие с доступом к финальной анкете после прохождения курса {course}:\n\nПользователь: {user_name}")
                    await self.vk_api.send_message(
                        user_id=user_id,
                        message=msg_template.format(course=text, user_name=user_name),
                        peer_id=peer_id,
                        keyboard=create_access_action_keyboard()
                    )
                    return
                
                # Handle "Открыть" - open access
                if session["step"] == "confirm_action" and text.startswith("🔓"):
                    course = session["selected_course"]
                    target_user_id = session["selected_user_id"]
                    field_name = f"access_survey_{course}"
                    
                    # Update database: 1 = accessible (show button)
                    await db.update_user_field(target_user_id, field_name, 1)
                    
                    # Get user info
                    target_user = await db.get_user(target_user_id)
                    user_name = target_user.get("user_name", "Unknown") if target_user else "Unknown"
                    
                    # Notify admin
                    msg_template = TEXTS_DATA.get("admin_access_opened", "✅ Доступ к финальной анкете курса {course} ОТКРЫТ для:\n{user_name}")
                    await self.vk_api.send_message(
                        user_id=user_id,
                        message=msg_template.format(course=course, user_name=user_name),
                        peer_id=peer_id,
                        keyboard=create_admin_keyboard()
                    )
                    
                    # Notify user
                    try:
                        msg_template = TEXTS_DATA.get("user_access_opened", "Вам открыт доступ к прохождению анкетирования после прохождения курса {course} - используйте кнопки меню")
                        await self.vk_api.send_message(
                            user_id=target_user_id,
                            message=msg_template.format(course=course),
                            peer_id=target_user_id
                        )
                    except Exception as e:
                        print(f"Failed to notify user {target_user_id}: {e}", flush=True)
                    
                    del ADMIN_SEARCH_SESSIONS[user_id]
                    return
                
                # Handle "Закрыть" - close access
                if session["step"] == "confirm_action" and text.startswith("🔒"):
                    course = session["selected_course"]
                    target_user_id = session["selected_user_id"]
                    field_name = f"access_survey_{course}"
                    
                    # Update database: 0 = inaccessible (hide button)
                    await db.update_user_field(target_user_id, field_name, 0)
                    
                    # Get user info
                    target_user = await db.get_user(target_user_id)
                    user_name = target_user.get("user_name", "Unknown") if target_user else "Unknown"
                    
                    # Notify admin
                    msg_template = TEXTS_DATA.get("admin_access_closed", "🔒 Доступ к финальной анкете курса {course} ЗАКРЫТ для:\n{user_name}")
                    await self.vk_api.send_message(
                        user_id=user_id,
                        message=msg_template.format(course=course, user_name=user_name),
                        peer_id=peer_id,
                        keyboard=create_admin_keyboard()
                    )
                    
                    del ADMIN_SEARCH_SESSIONS[user_id]
                    return
                
                # Handle search text input (only if not a menu command)
                menu_commands = ["меню", "админ", "менеджер", "маркетинг", "финальная анкета", "финальное анкетирование"]
                if session["step"] == "search" and text and text.lower() not in menu_commands:
                    # Search users by name
                    results = await db.search_users_by_name(text)
                    
                    if not results:
                        msg_template = TEXTS_DATA.get("admin_search_not_found", "По запросу \"{query}\" ничего не найдено.\n\nПопробуйте другой поиск:")
                        await self.vk_api.send_message(
                            user_id=user_id,
                            message=msg_template.format(query=text),
                            peer_id=peer_id,
                            keyboard=create_main_menu_keyboard()
                        )
                        return
                    
                    session["search_text"] = text
                    session["results"] = results
                    session["page"] = 0
                    session["step"] = "select_user"
                    
                    await self._show_search_results(user_id, peer_id)
                    return
            
            # Handle "Финальная анкета" / "Финальное анкетирование" button
            if text.lower() in ["финальная анкета", "финальное анкетирование"]:
                # Check access for the current group's course (access_survey_{N} == 1)
                course_index = self._get_course_index()
                user_data = await db.get_user(user_id)
                cf = self._get_user_course_fields(user_data, course_index)
                
                if cf["access_survey"] != 1:
                    await self.vk_api.send_message(
                        user_id=user_id,
                        message="У вас нет доступа к финальному анкетированию.",
                        peer_id=peer_id,
                        keyboard=create_main_menu_keyboard()
                    )
                    return
                
                await self._start_final_form(user_id, peer_id, course_index)
                return
            
            # Handle "Колесо фортуны" button from main menu
            if text.startswith("Колесо фортуны"):
                user_data = await db.get_user(user_id)
                fortune_wheel_spins = user_data.get("fortune_wheel", 0) if user_data else 0
                
                if fortune_wheel_spins <= 0:
                    await self.vk_api.send_message(
                        user_id=user_id,
                        message=TEXTS_DATA.get("fortune_wheel_no_spins", "У вас нет доступных вращений колеса фортуны."),
                        peer_id=peer_id,
                        keyboard=create_main_menu_keyboard()
                    )
                    return
                
                await self._handle_fortune_wheel_spin(user_id, peer_id)
                return
            
            # Handle fortune wheel buttons
            if text == "Крутить":
                await self._handle_fortune_wheel_spin(user_id, peer_id)
                return
            
            if text == "В другой раз":
                await self.vk_api.send_message(
                    user_id=user_id,
                    message=TEXTS_DATA.get("fortune_wheel_later", "Хорошо, вы сможете воспользоваться колесом фортуны позже через меню."),
                    peer_id=peer_id,
                    keyboard=create_main_menu_keyboard()
                )
                if user_id in FORTUNE_WHEEL_SESSIONS:
                    del FORTUNE_WHEEL_SESSIONS[user_id]
                return
            
            if text == "Испытать удачу":
                await self._spin_fortune_wheel(user_id, peer_id)
                return
            
            # Handle final form sessions (buttons, rating, open questions)
            if user_id in FINAL_FORM_SESSIONS:
                session = FINAL_FORM_SESSIONS[user_id]
                form_data = session.get("form_data", FINAL_FORM_DATA)
                question = get_final_form_question(session["current_question"], form_data)
                
                if question:
                    question_type = question.get("type", "open")
                    
                    # Handle button-type questions - only accept matching buttons
                    if question_type == "buttons":
                        buttons = question.get("buttons", [])
                        if text in buttons or text in ["Верно", "Изменить ФИ"]:
                            logger.info(f"Final form button: user={user_id}, qid={session['current_question']}, button={text}")
                            await self._handle_final_form_button(user_id, peer_id, text)
                            return
                        else:
                            # User typed text that doesn't match any button - ignore with message
                            print(f"User {user_id} typed '{text}' during button question (buttons: {buttons}), ignoring", flush=True)
                            await self.vk_api.send_message(
                                user_id=user_id,
                                message=TEXTS_DATA.get("use_buttons_message", "Пожалуйста, используйте кнопки для ответа."),
                                peer_id=peer_id
                            )
                            return
                    
                    # Handle rating questions (1-10)
                    if question_type == "rating":
                        try:
                            rating = int(text)
                            min_val = question.get("min", 1)
                            max_val = question.get("max", 10)
                            if min_val <= rating <= max_val:
                                logger.info(f"Final form rating: user={user_id}, qid={session['current_question']}, rating={rating}")
                                await self._handle_final_form_button(user_id, peer_id, text)
                                return
                            else:
                                # Rating out of range
                                print(f"User {user_id} typed rating {rating} out of range [{min_val}-{max_val}], ignoring", flush=True)
                                msg_template = TEXTS_DATA.get("rating_range_message", "Пожалуйста, введите число от {min} до {max}.")
                                await self.vk_api.send_message(
                                    user_id=user_id,
                                    message=msg_template.format(min=min_val, max=max_val),
                                    peer_id=peer_id
                                )
                                return
                        except ValueError:
                            # Not a number during rating question - ignore
                            print(f"User {user_id} typed '{text}' during rating question, ignoring", flush=True)
                            await self.vk_api.send_message(
                                user_id=user_id,
                                message=TEXTS_DATA.get("rating_number_message", "Пожалуйста, введите число от {min} до {max}.").format(
                                    min=question.get("min", 1), max=question.get("max", 10)
                                ),
                                peer_id=peer_id
                            )
                            return
                    
                    # Handle open questions (text input)
                    if question_type == "open":
                        logger.info(f"Final form answer: user={user_id}, qid={session['current_question']}, answer={text[:50]}")
                        await self._handle_final_form_answer(user_id, peer_id, text)
                        return
            
            # Handle "Сдал(а) практику" button
            if text.lower() == "сдал(а) практику" or text.lower() == "сдала практику" or text.lower() == "сдал практику":
                # Get user info for notification
                user_name = f"ID{user_id}"
                try:
                    user_info = await self.vk_api.get_user_info(user_id)
                    if "response" in user_info and user_info["response"]:
                        first_name = user_info["response"][0].get("first_name", "")
                        last_name = user_info["response"][0].get("last_name", "")
                        user_name = f"{first_name} {last_name}"
                except Exception as e:
                    print(f"Error getting user info: {e}", flush=True)
                
                # Update practice_{N} in database: 2 = completed (button pressed)
                course_index = self._get_course_index()
                await db.update_user_field(user_id, f"practice_{course_index}", 2)
                print(f"Practice {course_index} button clicked for user {user_id}, practice_{course_index} set to 2", flush=True)
                
                # Notify all managers (USER_MEN_IDS) - need to open access to final survey
                if USER_MEN_IDS:
                    user_link = f"[id{user_id}|{user_name}]"
                    msg_template = TEXTS_DATA.get("practice_notification", "🎯 Пользователь {user_link} сдал(а) Практику в Курс {course}!\n\nОткройте доступ к финальному анкетированию через АДМИН → Открыть доступ к Анкете")
                    admin_message = msg_template.format(user_link=user_link, course=course_index)
                    for admin_id in USER_MEN_IDS:
                        try:
                            await self.vk_api.send_message(
                                user_id=admin_id,
                                message=admin_message
                            )
                            print(f"Practice notification sent to manager {admin_id}", flush=True)
                        except Exception as e:
                            print(f"Failed to send practice notification to {admin_id}: {e}", flush=True)
                
                # Confirm to user - show main menu without practice button
                await self.vk_api.send_message(
                    user_id=user_id,
                    message=TEXTS_DATA.get("practice_confirmed", "✅ Уведомление о сдаче практики отправлено менеджеру.\n\nОжидайте, вам откроют доступ к финальному анкетированию."),
                    peer_id=peer_id,
                    keyboard=create_main_menu_keyboard()
                )
                return
            
            # Handle "Приветственная анкета" button
            if text.lower() in ["анкета", "приветственная анкета"]:
                # Check if user already completed form (form_first == 2)
                user_data = await db.get_user(user_id)
                form_first_status = user_data.get("form_first", 0) if user_data else 0
                
                if form_first_status == 2:
                    # Get all statuses for dynamic keyboard
                    course_index = self._get_course_index()
                    cf = self._get_user_course_fields(user_data, course_index)
                    fortune_wheel_spins = user_data.get("fortune_wheel", 0) if user_data else 0
                    
                    keyboard = create_dynamic_menu_keyboard(
                        is_admin=is_admin,
                    is_manager=is_manager,
                    is_marketing=is_marketing,
                        form_first=form_first_status,
                        test_book=cf["test_book"],
                        practice=cf["practice"],
                        access_survey=cf["access_survey"],
                        certificate=cf["diploma"],
                        fortune_wheel=fortune_wheel_spins,
                        course_index=course_index
                    )
                    await self.vk_api.send_message(
                        user_id=user_id,
                        message=TEXTS_DATA.get("form_already_completed", "Вы уже заполнили анкету!"),
                        peer_id=peer_id,
                        keyboard=keyboard
                    )
                    return
                
                # Send warning message
                warning_text = TEXTS_DATA.get("form_warning", "Если оборвать прохождение анкетирования, данные не сохранятся.")
                await self.vk_api.send_message(
                    user_id=user_id,
                    message=warning_text,
                    peer_id=peer_id,
                    keyboard=create_form_keyboard()
                )
                
                # Start form session
                FORM_SESSIONS[user_id] = {
                    "question": 0,
                    "answers": []
                }
                
                # Send start message
                start_text = TEXTS_DATA.get("form_start", "Начинаем заполнение анкеты.")
                await self.vk_api.send_message(
                    user_id=user_id,
                    message=start_text,
                    peer_id=peer_id
                )
                
                # Send first question
                await self._send_form_question(user_id, peer_id)
                return
            
            # Handle "Тестирование" button
            if text.lower() == "тестирование":
                # Check if user has access to test (test_book_{N} == 1)
                course_index = self._get_course_index()
                user_data = await db.get_user(user_id)
                cf = self._get_user_course_fields(user_data, course_index)
                
                if cf["test_book"] != 1:
                    # Get all statuses for dynamic keyboard
                    form_first_status = user_data.get("form_first", 0) if user_data else 0
                    fortune_wheel_spins = user_data.get("fortune_wheel", 0) if user_data else 0
                    
                    keyboard = create_dynamic_menu_keyboard(
                        is_admin=is_admin,
                    is_manager=is_manager,
                    is_marketing=is_marketing,
                        form_first=form_first_status,
                        test_book=cf["test_book"],
                        practice=cf["practice"],
                        access_survey=cf["access_survey"],
                        certificate=cf["diploma"],
                        fortune_wheel=fortune_wheel_spins,
                        course_index=course_index
                    )
                    await self.vk_api.send_message(
                        user_id=user_id,
                        message=TEXTS_DATA.get("test_not_available", "Сначала необходимо заполнить анкету!"),
                        peer_id=peer_id,
                        keyboard=keyboard
                    )
                    return
                
                await self._start_test(user_id, peer_id)
                return
            
            # Handle '📄 Скачать сертификат' button
            if text.lower() == "📄 скачать сертификат":
                # Check if user has certificate available for the current group's course
                course_index = self._get_course_index()
                user_data = await db.get_user(user_id)
                cf = self._get_user_course_fields(user_data, course_index)
                
                if cf["diploma"] != 1:
                    await self.vk_api.send_message(
                        user_id=user_id,
                        message="У вас нет доступного сертификата.",
                        peer_id=peer_id,
                        keyboard=create_main_menu_keyboard()
                    )
                    return
                
                # Generate and send certificate
                try:
                    user_name = user_data.get("user_name", "Участник") if user_data else "Участник"
                    
                    # Get current date
                    today = datetime.now()
                    date_str = today.strftime("%d.%m.%Y")
                    
                    # Generate certificate for current course (using only user_name)
                    certificate_data = generate_certificate(course_index, user_name, date_str)
                    if certificate_data:
                        filename = f"certificate_course_{course_index}_{user_id}.png"
                        await self.vk_api.send_document(
                            peer_id=peer_id,
                            file_data=certificate_data,
                            filename=filename,
                            message="🎓 Ваш сертификат:"
                        )
                    else:
                        await self.vk_api.send_message(
                            user_id=user_id,
                            message="Ошибка при генерации сертификата. Обратитесь к администратору.",
                            peer_id=peer_id,
                            keyboard=create_main_menu_keyboard()
                        )
                except Exception as e:
                    logger.error(f"Failed to send certificate to user {user_id}: {e}")
                    await self.vk_api.send_message(
                        user_id=user_id,
                        message="Ошибка при отправке сертификата. Обратитесь к администратору.",
                        peer_id=peer_id,
                        keyboard=create_main_menu_keyboard()
                    )
                return
            
            # Handle "Пройти заново" button
            restart_text = TEXTS_DATA.get("restart_button", "Пройти заново")
            if text.lower() == restart_text.lower():
                await self._start_test(user_id, peer_id)
                return
            
            # Handle test answer buttons (1, 2, 3) - only if in test session
            if text in ["1", "2", "3"] and user_id in USER_SESSIONS:
                await self._handle_answer(user_id, peer_id, int(text))
                return
            
            # Handle form answer (if user is in form session)
            if user_id in FORM_SESSIONS:
                await self._handle_form_answer(user_id, peer_id, text)
                return
            
            # Ignore all other text - send helpful message instead of silent ignore
            print(f"Ignoring text message from user {user_id}: {text}", flush=True)
            await self.vk_api.send_message(
                user_id=user_id,
                message=TEXTS_DATA.get("menu_select_action", "Выберите действие:"),
                peer_id=peer_id,
                keyboard=create_main_menu_keyboard()
            )
            
        except Exception as e:
            print(f"Message handling error: {e}", flush=True)

    async def _handle_download_db(self, user_id: int, peer_id: int) -> None:
        """Handle 'Скачать базу' button - export database to xlsx."""
        print(f"Download DB button pressed by user {user_id}", flush=True)
        
        try:
            # Get all users from database
            users = await db.get_all_users()
            
            if not users:
                await self.vk_api.send_message(
                    user_id=user_id,
                    message=TEXTS_DATA.get("db_empty", "База данных пуста."),
                    peer_id=peer_id,
                    keyboard=create_admin_keyboard()
                )
                return
            
            # Create XLSX file
            xlsx_data = create_users_xlsx(users)
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"users_{timestamp}.xlsx"
            
            # Send document
            msg_template = TEXTS_DATA.get("db_export_success", "📊 База данных: {count} пользователей")
            success = await self.vk_api.send_document(
                peer_id=peer_id,
                file_data=xlsx_data,
                filename=filename,
                message=msg_template.format(count=len(users))
            )
            
            if not success:
                await self.vk_api.send_message(
                    user_id=user_id,
                    message=TEXTS_DATA.get("db_export_error", "Ошибка при отправке файла."),
                    peer_id=peer_id,
                    keyboard=create_admin_keyboard()
                )
            
        except Exception as e:
            print(f"Error handling download DB: {e}", flush=True)
            msg_template = TEXTS_DATA.get("db_export_error_detail", "Ошибка: {error}")
            await self.vk_api.send_message(
                user_id=user_id,
                message=msg_template.format(error=e),
                peer_id=peer_id,
                keyboard=create_admin_keyboard()
            )

    async def _handle_import_db(self, user_id: int, peer_id: int, file_data: bytes, filename: str) -> None:
        """Handle Excel file import - update database from uploaded xlsx."""
        print(f"Import DB from file: {filename}", flush=True)
        
        try:
            # Parse xlsx file
            users = parse_users_xlsx(file_data)
            
            if not users:
                await self.vk_api.send_message(
                    user_id=user_id,
                    message="❌ Не удалось прочитать файл или файл пуст.",
                    peer_id=peer_id,
                    keyboard=create_admin_keyboard()
                )
                return
            
            print(f"Parsed {len(users)} users from file", flush=True)
            
            # Import users to database
            stats = await db.import_users(users)
            
            # Send result
            message = f"✅ Импорт завершён!\n\n"
            message += f"📊 Обработано записей: {len(users)}\n"
            message += f"➕ Создано новых: {stats['created']}\n"
            message += f"✏️ Обновлено: {stats['updated']}\n"
            if stats['errors'] > 0:
                message += f"⚠️ Ошибок: {stats['errors']}"
            
            await self.vk_api.send_message(
                user_id=user_id,
                message=message,
                peer_id=peer_id,
                keyboard=create_admin_keyboard()
            )
            
        except Exception as e:
            print(f"Error handling import DB: {e}", flush=True)
            await self.vk_api.send_message(
                user_id=user_id,
                message=f"❌ Ошибка при импорте: {e}",
                peer_id=peer_id,
                keyboard=create_admin_keyboard()
            )

    async def _start_test(self, user_id: int, peer_id: int) -> None:
        """Start new test for user."""
        course_index = self._get_course_index()
        td = get_tests_data(course_index)
        if not td or not td.get("variants"):
            print(f"ERROR: No test data for course {course_index}", flush=True)
            await self.vk_api.send_message(
                user_id=user_id,
                message="❌ Тесты для данного курса пока не загружены. Свяжитесь с администратором.",
                peer_id=peer_id,
                keyboard=create_main_menu_keyboard()
            )
            return
        
        # Send intro text with course-specific data
        intro_template = TEXTS_DATA.get("test_intro", "Начинаем тестирование!")
        total_q = td.get("test_info", {}).get("total_questions", 20)
        passing_sc = td.get("test_info", {}).get("passing_score", total_q - 2)
        intro_text = intro_template.format(
            course=course_index,
            passing_score=passing_sc,
            total_questions=total_q
        )
        await self.vk_api.send_message(
            user_id=user_id,
            message=intro_text,
            peer_id=peer_id
        )
        
        # Initialize session
        variant_idx = get_random_variant(course_index)
        USER_SESSIONS[user_id] = {
            "variant": variant_idx,
            "question": 0,
            "score": 0,
            "shuffled_answers": None,
            "course_index": course_index
        }
        
        # Send first question
        await self._send_question(user_id, peer_id)

    async def _send_question(self, user_id: int, peer_id: int) -> None:
        """Send current question to user."""
        session = USER_SESSIONS.get(user_id)
        if not session:
            print(f"No session for user {user_id}", flush=True)
            return
        
        course_index = session.get("course_index", 1)
        variant_idx = session["variant"]
        question_idx = session["question"]
        
        question = get_question(course_index, variant_idx, question_idx)
        if not question:
            print(f"No question found: variant={variant_idx}, question={question_idx}", flush=True)
            return
        
        # Shuffle answers
        shuffled = shuffle_answers(question.get("answers", []))
        session["shuffled_answers"] = shuffled
        
        # Get total questions
        td = get_tests_data(course_index)
        total = len(td.get("variants", [{}])[variant_idx].get("questions", []))
        
        # Format and send question
        question_text = format_question_message(question, question_idx + 1, total)
        
        # Add answer options
        answer_text = "\n\n1️⃣ " + shuffled[0]["text"]
        answer_text += "\n2️⃣ " + shuffled[1]["text"]
        answer_text += "\n3️⃣ " + shuffled[2]["text"]
        
        await self.vk_api.send_message(
            user_id=user_id,
            message=question_text + answer_text,
            peer_id=peer_id,
            keyboard=create_answer_keyboard(shuffled)
        )

    async def _handle_answer(self, user_id: int, peer_id: int, answer_num: int) -> None:
        """Handle user's answer."""
        session = USER_SESSIONS.get(user_id)
        if not session:
            print(f"No session for user {user_id}", flush=True)
            return
        
        shuffled = session.get("shuffled_answers", [])
        if answer_num < 1 or answer_num > len(shuffled):
            return
        
        # Check answer (answer_num is 1-indexed)
        selected_answer = shuffled[answer_num - 1]
        is_correct = selected_answer.get("is_correct", False)
        
        course_index = session.get("course_index", 1)
        variant_idx = session["variant"]
        question_idx = session["question"]
        question = get_question(course_index, variant_idx, question_idx)
        
        if is_correct:
            session["score"] += 1
            correct_text = TEXTS_DATA.get("correct_answer", "Верно!")
            await self.vk_api.send_message(
                user_id=user_id,
                message=correct_text,
                peer_id=peer_id
            )
        else:
            wrong_text = TEXTS_DATA.get("wrong_answer", "Неверно!")
            correct_answer = get_correct_answer_text(question)
            message = f"{wrong_text}\n\nПравильный ответ: {correct_answer}"
            await self.vk_api.send_message(
                user_id=user_id,
                message=message,
                peer_id=peer_id
            )
        
        # Move to next question
        session["question"] += 1
        
        # Check if test is complete
        td = get_tests_data(course_index)
        total = len(td.get("variants", [{}])[variant_idx].get("questions", []))
        
        if session["question"] >= total:
            await self._finish_test(user_id, peer_id)
        else:
            await self._send_question(user_id, peer_id)

    async def _finish_test(self, user_id: int, peer_id: int) -> None:
        """Finish test and show results."""
        session = USER_SESSIONS.get(user_id)
        if not session:
            return
        
        score = session["score"]
        course_index = session.get("course_index", 1)
        td = get_tests_data(course_index)
        total = len(td.get("variants", [{}])[session["variant"]].get("questions", []))
        passing_score = td.get("test_info", {}).get("passing_score", 18)
        passed = score >= passing_score
        
        # Check if user is admin/manager/marketing
        is_admin = (user_id == USER_ADMIN_ID)
        is_manager = (user_id in USER_MEN_IDS)
        is_marketing = (user_id in USER_MAR_IDS)
        
        # Get user name for notification
        user_name = f"ID{user_id}"
        user_display = f"[id{user_id}|ID{user_id}]"
        try:
            user_info = await self.vk_api.get_user_info(user_id)
            if "response" in user_info and user_info["response"]:
                first_name = user_info["response"][0].get("first_name", "")
                last_name = user_info["response"][0].get("last_name", "")
                user_name = f"{first_name} {last_name}"
                user_display = f"[id{user_id}|{user_name}]"
        except Exception as e:
            print(f"Error getting user info: {e}", flush=True)
        
        # Send notification to all managers
        # Use course_index from session (stored when test started)
        session = USER_SESSIONS.get(user_id)
        course_index = session.get("course_index", self._get_course_index()) if session else self._get_course_index()
        if USER_MEN_IDS:
            status = "СДАЛ" if passed else "НЕ СДАЛ"
            admin_message = f"Пользователь {user_display} {status} тест Курса {course_index}!\nРезультат: {score}/{total}"
            for men_id in USER_MEN_IDS:
                try:
                    await self.vk_api.send_message(
                        user_id=men_id,
                        message=admin_message
                    )
                    print(f"Manager notification sent to {men_id}", flush=True)
                except Exception as e:
                    print(f"Failed to send notification to {men_id}: {e}", flush=True)
        
        # Get user form status for correct keyboard
        user_data = await db.get_user(user_id)
        form_completed = user_data.get("form_first", False) if user_data else False
        has_final_survey_access = user_data.get(f"access_survey_{course_index}", False) if user_data else False
        
        # Send result to user
        if passed:
            # Update test_book_{N} in database: 2 = completed (test passed)
            await db.update_user_field(user_id, f"test_book_{course_index}", 2)
            
            # Set practice_{N} = 1 to show "Сдал(а) практику" button
            await db.update_user_field(user_id, f"practice_{course_index}", 1)
            
            # Get user data for keyboard
            user_data = await db.get_user(user_id)
            form_first_status = user_data.get("form_first", 0) if user_data else 0
            cf = self._get_user_course_fields(user_data, course_index)
            fortune_wheel_spins = user_data.get("fortune_wheel", 0) if user_data else 0
            
            keyboard = create_dynamic_menu_keyboard(
                is_admin=is_admin,
                    is_manager=is_manager,
                    is_marketing=is_marketing,
                form_first=form_first_status,
                test_book=cf["test_book"],
                practice=cf["practice"],
                access_survey=cf["access_survey"],
                certificate=cf["diploma"],
                fortune_wheel=fortune_wheel_spins,
                course_index=course_index
            )
            
            passed_text = TEXTS_DATA.get("test_passed", "🎉 Поздравляем! Вы сдали тест!")
            practice_info = "\n\nЕсли вы сдали практику - используйте кнопки Меню для уведомления менеджера. Он откроет вам доступ к финальному анкетированию и получению сертификата о прохождении курса."
            message = f"{passed_text}\n\nВаш результат: {score}/{total}{practice_info}"
            
            await self.vk_api.send_message(
                user_id=user_id,
                message=message,
                peer_id=peer_id,
                keyboard=keyboard
            )
        else:
            failed_text = TEXTS_DATA.get("test_failed", "К сожалению, вы не набрали нужное количество баллов.")
            message = f"{failed_text}\n\nВаш результат: {score}/{total}\nНеобходимо: {passing_score}/{total}"
            await self.vk_api.send_message(
                user_id=user_id,
                message=message,
                peer_id=peer_id,
                keyboard=create_retry_keyboard()
            )
        
        # Clear session
        del USER_SESSIONS[user_id]

    async def _send_form_question(self, user_id: int, peer_id: int) -> None:
        """Send current form question to user."""
        session = FORM_SESSIONS.get(user_id)
        if not session:
            print(f"No form session for user {user_id}", flush=True)
            return
        
        questions = FORM_DATA.get("questions", [])
        question_idx = session["question"]
        
        if question_idx >= len(questions):
            print(f"Question index out of range: {question_idx}", flush=True)
            return
        
        question = questions[question_idx]
        total = len(questions)
        
        # Format question message
        prefix_template = TEXTS_DATA.get("form_question_prefix", "Вопрос {current}/{total}:")
        prefix = prefix_template.format(current=question_idx + 1, total=total)
        message = f"{prefix}\n\n{question['question']}"
        
        await self.vk_api.send_message(
            user_id=user_id,
            message=message,
            peer_id=peer_id,
            keyboard=create_form_keyboard()
        )

    async def _handle_form_answer(self, user_id: int, peer_id: int, answer: str) -> None:
        """Handle user's answer to form question."""
        session = FORM_SESSIONS.get(user_id)
        if not session:
            print(f"No form session for user {user_id}", flush=True)
            return
        
        # Save answer
        session["answers"].append(answer)
        
        # Move to next question
        session["question"] += 1
        
        questions = FORM_DATA.get("questions", [])
        
        # Check if form is complete
        if session["question"] >= len(questions):
            await self._finish_form(user_id, peer_id)
        else:
            await self._send_form_question(user_id, peer_id)

    async def _finish_form(self, user_id: int, peer_id: int) -> None:
        """Finish form and save results."""
        session = FORM_SESSIONS.get(user_id)
        if not session:
            return
        
        answers = session["answers"]
        questions = FORM_DATA.get("questions", [])
        
        # Get user name from first answer (trimmed)
        user_name = answers[0].strip() if answers else f"ID{user_id}"
        
        # Format form answers as readable text
        form_answer_lines = []
        for i, (q, a) in enumerate(zip(questions, answers), 1):
            form_answer_lines.append(f"Вопрос {i}: {q['question']}")
            form_answer_lines.append(f"Ответ {i}: {a}")
            if i < len(questions):
                form_answer_lines.append("")  # Empty line between Q&A
        form_answer = "\n".join(form_answer_lines)
        
        # Update user in database: name, answers, form completed
        await db.update_user_field(user_id, "user_name", user_name)
        await db.update_user_field(user_id, "form_first_answer", form_answer)
        await db.update_user_field(user_id, "form_first", 2)  # 2 = completed
        # Unlock test access for ALL courses (welcome survey is shared)
        for _ci in range(1, 5):
            await db.update_user_field(user_id, f"test_book_{_ci}", 1)
        
        print(f"Form completed for user {user_id}, name: {user_name} (all courses unlocked)", flush=True)
        
        # Check if user is admin/manager/marketing
        is_admin = (user_id == USER_ADMIN_ID)
        is_manager = (user_id in USER_MEN_IDS)
        is_marketing = (user_id in USER_MAR_IDS)
        
        # Get user data for keyboard
        course_index = self._get_course_index()
        user_data = await db.get_user(user_id)
        form_first_status = user_data.get("form_first", 0) if user_data else 0
        cf = self._get_user_course_fields(user_data, course_index)
        fortune_wheel_spins = user_data.get("fortune_wheel", 0) if user_data else 0
        
        keyboard = create_dynamic_menu_keyboard(
            is_admin=is_admin,
            is_manager=is_manager,
            is_marketing=is_marketing,
            form_first=form_first_status,
            test_book=cf["test_book"],
            practice=cf["practice"],
            access_survey=cf["access_survey"],
            certificate=cf["diploma"],
            fortune_wheel=fortune_wheel_spins,
            course_index=course_index
        )
        
        # Send completion message
        completed_text = TEXTS_DATA.get("form_completed", "✅ Анкета успешно заполнена!")
        await self.vk_api.send_message(
            user_id=user_id,
            message=completed_text,
            peer_id=peer_id,
            keyboard=keyboard
        )
        
        # Send instructions
        instructions_text = TEXTS_DATA.get("instructions", "Инструкция")
        await self.vk_api.send_message(
            user_id=user_id,
            message=instructions_text,
            peer_id=peer_id,
            keyboard=keyboard
        )
        
        # Notify marketing (USER_MAR_IDS) about form completion with full answers
        if USER_MAR_IDS:
            # Create clickable link to user profile
            user_link = f"[id{user_id}|{user_name}]"
            admin_message = f"📋 Пользователь {user_link} заполнил приветственную анкету!\n\n{form_answer}"
            
            for mar_id in USER_MAR_IDS:
                try:
                    await self.vk_api.send_message(
                        user_id=mar_id,
                        message=admin_message
                    )
                    print(f"Marketing notification sent about form completion by user {user_id} to {mar_id}", flush=True)
                except Exception as e:
                    print(f"Failed to send notification to {mar_id}: {e}", flush=True)
        
        # Clear form session
        del FORM_SESSIONS[user_id]

    async def _show_search_results(self, user_id: int, peer_id: int) -> None:
        """Show search results with pagination."""
        session = ADMIN_SEARCH_SESSIONS.get(user_id)
        if not session:
            return
        
        results = session["results"]
        page = session["page"]
        per_page = 6
        total_pages = (len(results) + per_page - 1) // per_page if results else 1
        
        # Create keyboard with users
        keyboard = create_user_search_keyboard(results, page, per_page)
        
        # Create message
        search_text = session["search_text"]
        total = len(results)
        
        message = f"🔍 Результаты поиска \"{search_text}\":\n\n"
        message += f"Найдено: {total} пользователей\n"
        message += f"Страница {page + 1}/{total_pages}\n\n"
        message += "Выберите пользователя:"
        
        await self.vk_api.send_message(
            user_id=user_id,
            message=message,
            peer_id=peer_id,
            keyboard=keyboard
        )

    async def _start_final_form(self, user_id: int, peer_id: int, course: int) -> None:
        """Start final form for user."""
        # Check if user already completed this form (access_survey_X == 2)
        user_data = await db.get_user(user_id)
        access_field = f"access_survey_{course}"
        access_status = user_data.get(access_field, 0) if user_data else 0
        
        if access_status == 2:
            # Already completed
            await self.vk_api.send_message(
                user_id=user_id,
                message=TEXTS_DATA.get("final_form_already_completed", "Вы уже заполнили финальную анкету для этого курса!"),
                peer_id=peer_id,
                keyboard=create_main_menu_keyboard()
            )
            return
        
        if access_status != 1:
            # No access
            await self.vk_api.send_message(
                user_id=user_id,
                message="У вас нет доступа к финальному анкетированию.",
                peer_id=peer_id,
                keyboard=create_main_menu_keyboard()
            )
            return
        
        # Send warning
        warning_text = TEXTS_DATA.get("final_form_warning", "⚠️ Внимание! Если оборвать прохождение анкетирования, данные не сохранятся.")
        await self.vk_api.send_message(
            user_id=user_id,
            message=warning_text,
            peer_id=peer_id,
            keyboard=create_final_form_open_keyboard()
        )
        
        # Load form data for this course
        form_data = FINAL_FORMS_DATA.get(course)
        if not form_data:
            form_data = FINAL_FORM_DATA
            logger.warning(f"Form file for course {course} not found, using fallback (course 1)")
        
        # Get user name
        user_name = user_data.get("user_name", "") if user_data else ""
        
        # Store form_data in session for per-course form handling
        # Initialize session
        FINAL_FORM_SESSIONS[user_id] = {
            "course": course,
            "step": "question",
            "current_question": "start",  # Start with "Начать" button
            "answers": {},
            "user_name": user_name,
            "form_data": form_data
        }
        
        # Send first question
        await self._send_final_form_question(user_id, peer_id)

    async def _send_final_form_question(self, user_id: int, peer_id: int) -> None:
        """Send current final form question to user."""
        session = FINAL_FORM_SESSIONS.get(user_id)
        if not session:
            print(f"No final form session for user {user_id}", flush=True)
            return
        
        question_id = session["current_question"]
        form_data = session.get("form_data", FINAL_FORM_DATA)
        question = get_final_form_question(question_id, form_data)
        
        if not question:
            print(f"Question not found: {question_id}", flush=True)
            return
        
        question_text = question.get("question", "")
        question_type = question.get("type", "open")
        
        # Substitute variables in question text
        if "{user_name}" in question_text:
            question_text = question_text.replace("{user_name}", session.get("user_name", ""))
        
        # Send question text without prefix
        message = question_text
        
        # Choose keyboard based on question type
        if question_type == "open":
            keyboard = create_final_form_open_keyboard()
        elif question_type == "buttons":
            buttons = question.get("buttons", [])
            if buttons == ["Приступить"]:
                # Special keyboard for start button (final form)
                keyboard = {
                    "one_time": False,
                    "inline": False,
                    "buttons": [
                        [
                            {
                                "action": {"type": "text", "label": "Приступить"},
                                "color": "positive"
                            }
                        ]
                    ]
                }
            elif buttons == ["Да", "Нет"]:
                keyboard = create_yes_no_keyboard()
            elif "Не знаю" in buttons:
                keyboard = create_yes_no_unknown_keyboard()
            elif "Не против" in buttons:
                keyboard = create_agree_disagree_keyboard()
            elif "Верно" in buttons:
                keyboard = create_check_data_keyboard()
            else:
                keyboard = create_final_form_open_keyboard()
        elif question_type == "rating":
            keyboard = create_rating_keyboard(question.get("min", 1), question.get("max", 10))
        else:
            keyboard = create_final_form_open_keyboard()
        
        await self.vk_api.send_message(
            user_id=user_id,
            message=message,
            peer_id=peer_id,
            keyboard=keyboard
        )

    async def _handle_final_form_answer(self, user_id: int, peer_id: int, answer: str) -> None:
        """Handle answer in final form."""
        session = FINAL_FORM_SESSIONS.get(user_id)
        if not session:
            print(f"No final form session for user {user_id}", flush=True)
            return
        
        question_id = session["current_question"]
        form_data = session.get("form_data", FINAL_FORM_DATA)
        question = get_final_form_question(question_id, form_data)
        
        if not question:
            return
        
        question_type = question.get("type", "open")
        
        # Validate question type matches expected input (defense in depth)
        # If question is rating but came here as text — validate as rating
        if question_type == "rating":
            try:
                rating_val = int(answer.strip())
                min_val = question.get("min", 1)
                max_val = question.get("max", 10)
                if not (min_val <= rating_val <= max_val):
                    await self.vk_api.send_message(
                        user_id=user_id,
                        message=f"Пожалуйста, введите число от {min_val} до {max_val} или используйте кнопки.",
                        peer_id=peer_id,
                        keyboard=create_rating_keyboard(min_val, max_val)
                    )
                    return
                # Valid rating via text — delegate to button handler for proper branching
                await self._handle_final_form_button(user_id, peer_id, str(rating_val))
                return
            except ValueError:
                await self.vk_api.send_message(
                    user_id=user_id,
                    message="Пожалуйста, используйте кнопки для оценки или введите число.",
                    peer_id=peer_id,
                    keyboard=create_rating_keyboard(question.get("min", 1), question.get("max", 10))
                )
                return
        
        # If question is buttons but came here as text — reject
        elif question_type == "buttons":
            await self.vk_api.send_message(
                user_id=user_id,
                message=TEXTS_DATA.get("use_buttons_message", "Пожалуйста, используйте кнопки для ответа."),
                peer_id=peer_id
            )
            return
        
        # For "open" questions — accept any text
        # Save answer using question_id as key (for proper formatting in finish)
        session["answers"][question_id] = answer
        
        # Check if this question updates user fields
        update_field = question.get("update_field")
        if update_field:
            if not answer or len(answer.strip()) > 200:
                await self._send_message(peer_id, "Введите корректные данные (до 200 символов)")
                return
            await db.update_user_field(user_id, update_field, answer.strip())
            if update_field == "user_name":
                session["user_name"] = answer.strip()
                # After editing name, go back to verify_name
                session["current_question"] = "verify_name"
                await self._send_final_form_question(user_id, peer_id)
                return
        
        # Get next question
        next_question = get_next_question_id(question_id, answer, form_data)
        
        if next_question == "finish" or next_question is None:
            await self._finish_final_form(user_id, peer_id)
            return
        else:
            session["current_question"] = next_question
        
        await self._send_final_form_question(user_id, peer_id)

    async def _handle_final_form_button(self, user_id: int, peer_id: int, button: str) -> None:
        """Handle button press in final form."""
        session = FINAL_FORM_SESSIONS.get(user_id)
        if not session:
            return
        
        question_id = session["current_question"]
        form_data = session.get("form_data", FINAL_FORM_DATA)
        question = get_final_form_question(question_id, form_data)
        
        if not question:
            return
        
        question_type = question.get("type", "open")
        
        # For rating questions, validate the button value
        if question_type == "rating":
            try:
                rating = int(button)
                min_val = question.get("min", 1)
                max_val = question.get("max", 10)
                if rating < min_val or rating > max_val:
                    print(f"User {user_id} sent rating {rating} out of range [{min_val}-{max_val}], ignoring", flush=True)
                    await self.vk_api.send_message(
                        user_id=user_id,
                        message=f"Пожалуйста, введите число от {min_val} до {max_val}.",
                        peer_id=peer_id,
                        keyboard=create_rating_keyboard(min_val, max_val)
                    )
                    return
            except ValueError:
                print(f"User {user_id} sent non-number '{button}' for rating question, ignoring", flush=True)
                return
        
        # For button questions, validate that button is in allowed list
        elif question_type == "buttons":
            allowed_buttons = question.get("buttons", [])
            if button not in allowed_buttons:
                print(f"Unexpected button '{button}' for question {question_id} (allowed: {allowed_buttons})", flush=True)
                return
        
        # Save answer using question_id as key (skip "start" button - not a real answer)
        if question_id != "start":
            session["answers"][question_id] = button
        
        # Handle verify_name (verification of user_name for certificate)
        if str(question_id) == "verify_name":
            if button == "Верно":
                await self._finish_final_form(user_id, peer_id)
                return
            elif button == "Изменить ФИ":
                session["current_question"] = "edit_name"
                await self._send_final_form_question(user_id, peer_id)
                return
        
        # Get next question
        next_question = get_next_question_id(question_id, button, form_data)
        
        if next_question == "finish" or next_question is None:
            await self._finish_final_form(user_id, peer_id)
            return
        
        session["current_question"] = next_question
        await self._send_final_form_question(user_id, peer_id)

    async def _finish_final_form(self, user_id: int, peer_id: int) -> None:
        """Finish final form and save results."""
        session = FINAL_FORM_SESSIONS.get(user_id)
        if not session:
            return
        
        course = session["course"]
        answers = session["answers"]  # Format: {question_id: answer}
        
        # Build question texts dictionary for formatting
        form_data = session.get("form_data", FINAL_FORM_DATA)
        question_texts = {}
        for q in form_data.get("questions", []):
            qid = q.get("id")
            qtext = q.get("question", "")
            # Remove variable placeholders for storage
            qtext_clean = qtext.replace("{user_name}", "").strip()
            question_texts[qid] = qtext_clean
        
        # Format answers with question text (same format as initial form)
        # Sort by question id - numeric first, then alphanumeric
        def sort_key(qid):
            if isinstance(qid, int):
                return (0, qid)
            try:
                # Try to extract numeric part (e.g., "2a" -> 2.1)
                num_part = ""
                for c in str(qid):
                    if c.isdigit():
                        num_part += c
                    else:
                        break
                if num_part:
                    return (0, int(num_part))
                return (1, str(qid))
            except:
                return (1, str(qid))
        
        sorted_qids = sorted(answers.keys(), key=sort_key)
        
        answers_text_lines = []
        question_num = 0
        for qid in sorted_qids:
            answer = answers[qid]
            qtext = question_texts.get(qid, f"Вопрос {qid}")
            question_num += 1
            answers_text_lines.append(f"Вопрос {question_num}: {qtext}")
            answers_text_lines.append(f"Ответ {question_num}: {answer}")
            if question_num < len(sorted_qids):
                answers_text_lines.append("")  # Empty line between Q&A
        
        answers_text = "\n".join(answers_text_lines)
        
        # Save to database
        form_field = f"form_end_{course}"
        await db.update_user_field(user_id, form_field, answers_text)
        
        # Update status: access_survey_X = 2 (completed)
        access_field = f"access_survey_{course}"
        await db.update_user_field(user_id, access_field, 2)
        
        # Make certificate accessible: diploma_X = 1
        certificate_field = f"diploma_{course}"
        await db.update_user_field(user_id, certificate_field, 1)
        
        print(f"Final form completed for user {user_id}, course {course}", flush=True)
        
        # Check if user is admin/manager/marketing
        is_admin = (user_id == USER_ADMIN_ID)
        is_manager = user_id in USER_MEN_IDS
        is_marketing = user_id in USER_MAR_IDS
        
        # Get user data for keyboard (use current course index)
        course_index = self._get_course_index()
        user_data = await db.get_user(user_id)
        form_first_status = user_data.get("form_first", 0) if user_data else 0
        cf = self._get_user_course_fields(user_data, course_index)
        fortune_wheel_spins = user_data.get("fortune_wheel", 0) if user_data else 0
        
        keyboard = create_dynamic_menu_keyboard(
            is_admin=is_admin,
            is_manager=is_manager,
            is_marketing=is_marketing,
            form_first=form_first_status,
            test_book=cf["test_book"],
            practice=cf["practice"],
            access_survey=cf["access_survey"],
            certificate=cf["diploma"],
            fortune_wheel=fortune_wheel_spins,
            course_index=course_index
        )
        
        # Send congratulations
        congrats_text = TEXTS_DATA.get("diploma_congratulations", "Поздравляем с окончанием курса💥\nС радостью вручаем сертификат🥳")
        await self.vk_api.send_message(
            user_id=user_id,
            message=congrats_text,
            peer_id=peer_id,
            keyboard=keyboard
        )
        
        # Generate and send certificate
        try:
            user_data = await db.get_user(user_id)
            user_name = session.get("user_name", "").strip() or (user_data.get("user_name", "") if user_data else "")
            if not user_name:
                user_name = "Участник"
            
            # Get current date
            today = datetime.now()
            date_str = today.strftime("%d.%m.%Y")
            
            # Generate certificate (using only user_name, NOT komu_vydan)
            certificate_data = generate_certificate(course, user_name, date_str)
            if certificate_data:
                filename = f"certificate_course_{course}_{user_id}.png"
                await self.vk_api.send_document(
                    peer_id=peer_id,
                    file_data=certificate_data,
                    filename=filename,
                    message="🎓 Ваш сертификат:"
                )
        except Exception as e:
            logger.error(f"Failed to send certificate to user {user_id}: {e}")
        
        # Notify marketing (USER_MAR_IDS) about final form completion
        if USER_MAR_IDS:
            user_data = await db.get_user(user_id)
            user_name = user_data.get("user_name", f"ID{user_id}") if user_data else f"ID{user_id}"
            user_link = f"[id{user_id}|{user_name}]"
            msg_template = TEXTS_DATA.get("final_form_admin_notification", "🎓 Пользователь {user_link} завершил Курс {course}!\n\n{form_answers}")
            mar_message = msg_template.format(user_link=user_link, course=course, form_answers=answers_text)
            for mar_id in USER_MAR_IDS:
                try:
                    await self.vk_api.send_message(
                        user_id=mar_id,
                        message=mar_message
                    )
                except Exception as e:
                    print(f"Failed to notify marketing {mar_id}: {e}", flush=True)
        
        # Clear session
        del FINAL_FORM_SESSIONS[user_id]
        
        # Add fortune wheel spin
        await db.increment_fortune_wheel(user_id, 1)
        
        # Offer fortune wheel
        await self._offer_fortune_wheel(user_id, peer_id)

    async def _offer_fortune_wheel(self, user_id: int, peer_id: int) -> None:
        """Offer fortune wheel to user after final form completion."""
        user_data = await db.get_user(user_id)
        fortune_wheel_spins = user_data.get("fortune_wheel", 0) if user_data else 0
        
        if fortune_wheel_spins <= 0:
            return
        
        # Offer wheel
        intro_text = TEXTS_DATA.get("fortune_wheel_intro", "Вам доступно вращение \"Колеса фортуны\"")
        await self.vk_api.send_message(
            user_id=user_id,
            message=intro_text,
            peer_id=peer_id,
            keyboard=create_fortune_wheel_keyboard()
        )

    async def _handle_fortune_wheel_spin(self, user_id: int, peer_id: int) -> None:
        """Handle fortune wheel spin request - show prizes."""
        user_data = await db.get_user(user_id)
        fortune_wheel_spins = user_data.get("fortune_wheel", 0) if user_data else 0
        
        if fortune_wheel_spins <= 0:
            await self.vk_api.send_message(
                user_id=user_id,
                message=TEXTS_DATA.get("fortune_wheel_no_spins", "У вас нет доступных вращений колеса фортуны."),
                peer_id=peer_id,
                keyboard=create_main_menu_keyboard()
            )
            return
        
        # Show prizes and spin button
        prizes_list = get_sorted_prizes_list()
        ready_text = TEXTS_DATA.get("fortune_wheel_ready", "Готовы испытать удачу?\nПризы, которые мы разыгрываем:\n\n{prizes_list}")
        await self.vk_api.send_message(
            user_id=user_id,
            message=ready_text.format(prizes_list=prizes_list),
            peer_id=peer_id,
            keyboard=create_spin_wheel_keyboard()
        )

    async def _spin_fortune_wheel(self, user_id: int, peer_id: int) -> None:
        """Spin the fortune wheel and award prize."""
        course_index = self._get_course_index()
        user_data = await db.get_user(user_id)
        fortune_wheel_spins = user_data.get("fortune_wheel", 0) if user_data else 0
        
        if fortune_wheel_spins <= 0:
            # Get user state for dynamic keyboard
            is_admin = (user_id == USER_ADMIN_ID)
            is_manager = user_id in USER_MEN_IDS
            is_marketing = user_id in USER_MAR_IDS
            form_first_status = user_data.get("form_first", 0) if user_data else 0
            cf = self._get_user_course_fields(user_data, course_index)
            
            await self.vk_api.send_message(
                user_id=user_id,
                message=TEXTS_DATA.get("fortune_wheel_no_spins", "У вас нет доступных вращений колеса фортуны."),
                peer_id=peer_id,
                keyboard=create_dynamic_menu_keyboard(
                    is_admin=is_admin,
                    is_manager=is_manager,
                    is_marketing=is_marketing,
                    form_first=form_first_status,
                    test_book=cf["test_book"],
                    practice=cf["practice"],
                    access_survey=cf["access_survey"],
                    certificate=cf["diploma"],
                    fortune_wheel=fortune_wheel_spins,
                    course_index=course_index
                )
            )
            return
        
        # Select prize
        prizes = FINAL_FORM_DATA.get("fortune_wheel", {}).get("prizes", [])
        prize = select_prize_by_probability(prizes)
        
        # Decrease fortune wheel spins
        await db.increment_fortune_wheel(user_id, -1)
        
        # Get updated user data for keyboard
        user_data = await db.get_user(user_id)
        fortune_wheel_spins = user_data.get("fortune_wheel", 0) if user_data else 0
        
        is_admin = (user_id == USER_ADMIN_ID)
        is_manager = user_id in USER_MEN_IDS
        is_marketing = user_id in USER_MAR_IDS
        form_first_status = user_data.get("form_first", 0) if user_data else 0
        cf = self._get_user_course_fields(user_data, course_index)
        
        keyboard = create_dynamic_menu_keyboard(
            is_admin=is_admin,
            is_manager=is_manager,
            is_marketing=is_marketing,
            form_first=form_first_status,
            test_book=cf["test_book"],
            practice=cf["practice"],
            access_survey=cf["access_survey"],
            certificate=cf["diploma"],
            fortune_wheel=fortune_wheel_spins,
            course_index=course_index
        )
        
        # Notify about prize
        result_text = TEXTS_DATA.get("fortune_wheel_result", "🎉 Поздравляем! Вам выпало:\n\n🎁 {prize}")
        await self.vk_api.send_message(
            user_id=user_id,
            message=result_text.format(prize=prize),
            peer_id=peer_id,
            keyboard=keyboard
        )
        
        # Send instructions for getting prize
        if is_physical_prize(prize):
            instruction_text = TEXTS_DATA.get("prize_physical", "Для того, что бы получить свой приз, подойдите к тренеру перед тренировкой и покажите скрин экрана с выпавшим призом!")
        else:
            instruction_text = TEXTS_DATA.get("prize_digital", "С Вами в ближайшее время свяжется персональный менеджер и расскажет, как забрать свой приз!")
        
        await self.vk_api.send_message(
            user_id=user_id,
            message=instruction_text,
            peer_id=peer_id,
            keyboard=keyboard
        )
        
        # Notify managers about digital (non-physical) prize
        if not is_physical_prize(prize) and USER_MEN_IDS:
            user_data = await db.get_user(user_id)
            user_name = user_data.get("user_name", f"ID{user_id}") if user_data else f"ID{user_id}"
            user_link = f"[id{user_id}|{user_name}]"
            prize_notification = f"🎁 Пользователь {user_link} выиграл цифровой приз: {prize}\n\nСвяжитесь с пользователем для выдачи приза."
            for men_id in USER_MEN_IDS:
                try:
                    await self.vk_api.send_message(
                        user_id=men_id,
                        message=prize_notification
                    )
                except Exception as e:
                    print(f"Failed to notify manager {men_id} about prize: {e}", flush=True)


# -----------------------------------------------------------------------------
# Entrypoint
# -----------------------------------------------------------------------------

# TODO: ПРИ СЛЕДУЮЩЕМ ОБРАЩЕНИИ К АГЕНТУ — УДАЛИТЬ ФУНКЦИЮ run_db_column_cleanup() И ЕЁ ВЫЗОВ.
# TODO: [МИГРАЦИЯ] УДАЛИТЬ ЭТУ ФУНКЦИЮ ПОСЛЕ УСПЕШНОГО ВЫПОЛНЕНИЯ
async def run_db_column_cleanup():
    """Удаляет устаревшие столбцы, если они существуют."""
    columns_to_drop = ["komu_vydan", "user_name_case", "ФИ_Датапд"]
    logger.info("🔄 Запуск очистки устаревших столбцов БД...")
    
    async with db.pool.acquire() as conn:
        for col in columns_to_drop:
            try:
                await conn.execute(f'''
                    ALTER TABLE users 
                    DROP COLUMN IF EXISTS "{col}"
                ''')
                logger.info(f"✅ Столбец '{col}' успешно удалён или отсутствует.")
            except Exception as e:
                logger.warning(f"⚠️ Ошибка при удалении столбца '{col}': {e}")
    logger.info("✅ Миграция очистки БД завершена.")


async def async_main():
    """Async main entrypoint."""
    print("Creating web server...", flush=True)
    server = WebServer()
    
    # Initialize database
    await db.init()
    
    # TODO: Раскомментировать для одноразового запуска миграции, затем удалить
    # if os.getenv("RUN_MIGRATION") == "1":
    #     await run_db_column_cleanup()
    
    runner = web.AppRunner(server.app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", PORT)
    await site.start()
    
    print(f"Server started on port {PORT}", flush=True)
    
    try:
        await asyncio.Event().wait()
    except KeyboardInterrupt:
        print("Shutting down...", flush=True)
    finally:
        await db.close()
        await runner.cleanup()
        print("Application shutdown complete", flush=True)


if __name__ == "__main__":
    asyncio.run(async_main())
